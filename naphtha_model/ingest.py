"""Ingest the desk's 'Estimated Refinery Outputs' workbook (2024 yields).

    python -m naphtha_model ingest-yields <file.xlsx>

Does two things:
1. Writes data/reference/refinery_yields_2024.csv — per-refinery 2024 product
   yields (% of crude), naphtha included. This is NET merchant naphtha
   (after reformer/isom pull), matching the model's net-naphtha concept.
2. Merges every refinery into data/reference/refineries.csv. Existing rows
   (matched by city + PADD + operator token) keep their id, name, capacity,
   units and notes; new rows arrive with crude_capacity_kbd=0 ("yield-mode")
   until the capacity sheet lands.

Refineries that have no unit detail get a synthetic CRUDE-EST unit at load
time (see loaders.py): net naphtha = crude capacity x utilization x 2024
yield, split across cuts by assumptions/global.yaml yield_mode.cut_shares.
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path

import openpyxl

from .config import DATA_DIR


def _read_utf16_tsv(path: Path) -> list[list[str]]:
    """REM / RefineryDataTool exports are UTF-16 tab-separated."""
    text = Path(path).open(encoding="utf-16").read()
    return [
        [c.strip() for c in row]
        for row in csv.reader(io.StringIO(text), delimiter="\t")
    ]


def _num(s: str) -> float:
    s = (s or "").replace(",", "").replace("%", "").strip()
    return float(s) if s else 0.0

ROMAN_PADD = {"PADD I": 1, "PADD II": 2, "PADD III": 3, "PADD IV": 4, "PADD V": 5}

YIELD_COLS = [
    ("gasoil_diesel_pct", 8),
    ("gasoline_pct", 9),
    ("hfo_pct", 10),
    ("kero_jet_pct", 11),
    ("lpg_pct", 12),
    ("naphtha_pct", 13),
    ("total_pct", 14),
]


def _blank(v) -> bool:
    return v is None or str(v).strip() == ""


def _slug(operator: str, city: str) -> str:
    op_token = re.sub(r"[^A-Za-z0-9]", "", operator.split()[0]).upper()
    city_token = re.sub(r"[^A-Za-z0-9]+", "_", city.strip()).upper().strip("_")
    return f"{op_token}_{city_token}"


def parse_yields_workbook(path: Path) -> list[dict]:
    """One dict per refinery. Handles the pivot-style layout: region/country/
    PADD/state/city forward-filled, extra owner rows (blank operator) folded
    into the owner string of the refinery above."""
    ws = openpyxl.load_workbook(path, data_only=True).active
    refineries: list[dict] = []
    cur = [None] * 5  # region, country, padd, state, city
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "Region" or all(_blank(v) for v in row):
            continue
        for i in range(5):
            if not _blank(row[i]):
                cur[i] = str(row[i]).strip()
        operator = row[5]
        if _blank(operator):
            # extra owner row for the refinery above (JV)
            if not _blank(row[6]) and refineries and str(row[6]).strip() != "Owner":
                refineries[-1]["owners"].append((str(row[6]).strip(), row[7]))
            continue
        if cur[2] not in ROMAN_PADD:
            raise ValueError(f"unrecognised PADD {cur[2]!r} for {operator}")
        rec = {
            "padd": ROMAN_PADD[cur[2]],
            "state": cur[3] or "",
            "city": cur[4] or "",
            "operator": str(operator).strip(),
            "owners": [(str(row[6]).strip(), row[7])] if not _blank(row[6]) else [],
        }
        for name, col in YIELD_COLS:
            rec[name] = round(float(row[col] or 0.0), 4)
        refineries.append(rec)
    return refineries


def _owner_string(owners: list[tuple[str, object]]) -> str:
    if not owners:
        return ""
    if len(owners) == 1 and float(owners[0][1] or 100) == 100:
        return owners[0][0]
    return " / ".join(f"{name} ({float(share or 0):.0f}%)" for name, share in owners)


def _read_registry(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def _match_existing(rec: dict, registry: list[dict]) -> dict | None:
    """Same city + PADD, and the operator's first token appears in the
    existing name or owner — keeps hand-curated ids like MOTIVA_PAR."""
    token = rec["operator"].split()[0].lower()
    for row in registry:
        if (
            row["city"].strip().lower() == rec["city"].strip().lower()
            and int(row["padd"]) == rec["padd"]
            and (token in row["name"].lower() or token in row["owner"].lower())
        ):
            return row
    return None


def ingest_yields(xlsx_path: Path, data_dir: Path = DATA_DIR) -> dict:
    records = parse_yields_workbook(Path(xlsx_path))
    registry_path = data_dir / "reference" / "refineries.csv"
    registry = _read_registry(registry_path)
    # match only against rows that existed before this ingestion, and let
    # each pre-existing row be claimed at most once
    preexisting = list(registry)
    claimed: set[str] = set()
    matched = added = 0
    used_ids = {row["refinery_id"] for row in registry}
    yields_rows = []

    for rec in records:
        existing = _match_existing(rec, preexisting)
        if existing is not None and existing["refinery_id"] in claimed:
            existing = None
        if existing is not None:
            claimed.add(existing["refinery_id"])
            rid = existing["refinery_id"]
            matched += 1
        else:
            rid = _slug(rec["operator"], rec["city"])
            base, n = rid, 2
            while rid in used_ids:
                rid = f"{base}_{n}"
                n += 1
            used_ids.add(rid)
            registry.append(
                {
                    "refinery_id": rid,
                    "name": f"{rec['operator']} {rec['city']}",
                    "owner": _owner_string(rec["owners"]) or rec["operator"],
                    "city": rec["city"],
                    "state": rec["state"],
                    "padd": str(rec["padd"]),
                    "region": "US",
                    "crude_capacity_kbd": "0",
                    "status": "operating",
                    "notes": "yield-mode; awaiting capacity sheet",
                }
            )
            added += 1
        yields_rows.append(
            {
                "refinery_id": rid,
                "padd": rec["padd"],
                "state": rec["state"],
                "city": rec["city"],
                "operator": rec["operator"],
                "owners": _owner_string(rec["owners"]),
                **{name: rec[name] for name, _ in YIELD_COLS},
            }
        )

    reg_fields = ["refinery_id", "name", "owner", "city", "state", "padd",
                  "region", "crude_capacity_kbd", "status", "notes"]
    with registry_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=reg_fields)
        w.writeheader()
        for row in registry:
            w.writerow({k: row.get(k, "") for k in reg_fields})

    yields_path = data_dir / "reference" / "refinery_yields_2024.csv"
    y_fields = ["refinery_id", "padd", "state", "city", "operator", "owners"] + [
        name for name, _ in YIELD_COLS
    ]
    with yields_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=y_fields)
        w.writeheader()
        w.writerows(yields_rows)

    return {"refineries": len(records), "matched_existing": matched,
            "added_to_registry": added, "yields_csv": str(yields_path),
            "registry_csv": str(registry_path)}


# ------------------------------------------------------------------ capacity
#
# EA site-level monthly nameplate capacity (kb/d), 2023-2026. Site names are
# city-based with parenthetical disambiguators; the alias map below resolves
# every site that automatic matching (exact city, then operator hint, then
# name-contains) cannot. None = site is not in the registry (shut before the
# 2024 yields vintage, or a splitter/topping asset not yet modelled) — its
# monthly series is still written to site_capacity_monthly.csv for reference.

SITE_ALIASES: dict[str, str | None] = {
    "Anacortes": "MARATHON_ANACORTES",            # ex-Tesoro/Andeavor plant
    "Anacortes (Puget Sound)": "HF_ANACORTES",    # ex-Shell Puget Sound
    "Bakersfield": None,                          # ex-Big West, converted/shut
    "Bakersfield (SJR)": "SAN_BAKERSFIELD",       # San Joaquin Refining
    "Bayway": "PHILLIPS_LINDEN",                  # P66 Bayway, Linden NJ
    "Billings": "PAR_BILLINGS",                   # ex-ExxonMobil Billings
    "Cherry Point": "BP_FERNDALE",                # BP Cherry Point
    "Corpus Christi (Bill Greehey)": "VALERO_CORPUS_CHRISTI",
    "Corpus Christi (Trigeant)": None,            # asphalt, not in registry
    "Corpus Christi Splitter (Magellan)": None,   # splitter, not in registry yet
    "Channelview Splitter": None,                 # splitter, not in registry yet
    "Eagle Springs (Ely)": "FORELAND_TONOPAH",    # Foreland, NV
    "Ferndale": "PHILLIPS_FERNDALE",
    "Galena Park": "KINDER_HOUSTON",              # Kinder Morgan splitters
    "Galveston Bay": "MPC_GALV_BAY",
    "Houston Splitter": None,                     # splitter, not in registry yet
    "Kapolei (Ewa Beach -East Plant)": "PAR_EWA_BEACH",
    "Kenai": "MARATHON_NIKISKI",
    "Kuparuk Topping Unit": "CONOCOPHILLIPS_ANCHORAGE",
    "Lake Charles (Pelican)": None,
    "Lake Charles (Westlake - P66)": "PHILLIPS_WESTLAKE",
    "Los Angeles (Wilmington and Carson)": "PHILLIPS_LOS_ANGELES",  # shut 2025
    "McKee (Sunray)": "VALERO_SUNRAY",
    "Navajo (Artesia)": "HF_ARTESIA",
    "North Pole (FHR)": None,                     # FHR North Pole, shut
    "North Salt Lake": "BIG_NORTH_SLATERVILLE_CANAL",  # Big West Oil
    "Paulsboro": None,                            # asphalt plant, not PBF
    "Paulsboro (PBF)": "PBF_PAULSBORO",
    "Pine Bend (Rosemount)": "FLINT_ROSEMOUNT",
    "Port Arthur (Total)": "TOTAL_PORT_ARTHUR",
    "Saraland": "VERTEX_MOBILE",                  # Vertex Saraland (Mobile AL)
    "Sinclair, WY": "SINCLAIR_SINCLAIR",
    "Wilmington (Marathon)": "MARATHON_LOS_ANGELES",
    "Wilmington (Valero)": "VALERO_LOS_ANGELES",
    "Wilmington Asphalt Refinery": None,
    "Woods Cross": "HF_WOODS_CROSS",
    # shut sites with no registry row (pre-2024 closures):
    "Alliance (Belle Chasse)": None, "Bloomfield": None, "Brownsville": None,
    "Cheyenne": None, "Davis": None, "Dickinson": None,
    "Gallup": None, "Golden Eagle (Martinez)": None, "Paramount": None,
    "Par West (Barbers Point, ex-Island Energy)": None, "Perth Amboy": None,
    "Philadelphia Complex": None, "Rodeo": None, "Santa Maria": None,
    "Santa Maria (Arroyo Grande)": None, "Savannah": None, "Southland": None,
}

# registry rows that are duplicate/component entities in the EA yields data;
# their capacity is carried on the primary row
COMPONENT_NOTES = {
    "MARATHON_TEXAS_CITY_GALVESTON_BAY":
        "EA yields component of the Galveston Bay complex; "
        "capacity carried on MPC_GALV_BAY",
}

_STATE_SUFFIX = re.compile(
    r",\s*(mt|wv|tx|la|ca|wy|ut|nd|mn|ks|ok|pa|nj|wa|ak|hi|nm|co|nv|mi|oh|il|in|"
    r"ky|tn|wi|al|ar|ms|de|ga|sc|va|md|mo)$"
)

_HINT_ALIASES = {"p66": "phillips", "cvx": "chevron", "mpc": "marathon",
                 "fhr": "flint", "basf total": "basf"}


def _norm(s: str) -> str:
    s = _STATE_SUFFIX.sub("", s.lower().strip())
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def match_site(site: str, registry: list[dict]) -> str | None:
    """Resolve an EA site name to a registry refinery_id (None = no match)."""
    if site in SITE_ALIASES:
        return SITE_ALIASES[site]
    m = re.match(r"^(.*?)\s*(?:\((.*)\))?$", site)
    base, hint = _norm(m.group(1)), _norm(m.group(2) or "")
    hint = _HINT_ALIASES.get(hint, hint)

    exact = [r for r in registry if _norm(r["city"]) == base]
    if len(exact) == 1:
        return exact[0]["refinery_id"]
    cands = exact or [
        r for r in registry
        if base in _norm(r["name"]) or (_norm(r["city"]) and _norm(r["city"]) in base)
    ]
    if len(cands) > 1 and hint:
        hinted = [r for r in cands
                  if hint.split()[0] in _norm(r["name"] + " " + r["owner"])]
        if len(hinted) == 1:
            return hinted[0]["refinery_id"]
        cands = hinted or cands
    if len(cands) == 1:
        return cands[0]["refinery_id"]
    return None


def ingest_capacity(
    csv_path: Path, data_dir: Path = DATA_DIR, as_of: str = "2026-07"
) -> dict:
    """Load EA site nameplate capacities: stamp the as-of month into the
    registry (multi-site refineries are summed), write the full monthly
    series to data/reference/site_capacity_monthly.csv."""
    registry_path = data_dir / "reference" / "refineries.csv"
    registry = _read_registry(registry_path)

    with Path(csv_path).open(newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))

    sites = sorted({r["refinery"] for r in rows})
    site_to_rid = {s: match_site(s, registry) for s in sites}

    monthly_path = data_dir / "reference" / "site_capacity_monthly.csv"
    with monthly_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["site", "refinery_id", "month", "capacity_kbd"])
        for r in sorted(rows, key=lambda r: (r["refinery"], r["date"])):
            w.writerow([r["refinery"], site_to_rid[r["refinery"]] or "",
                        r["date"][:7], r["value"]])

    # as-of capacity per refinery (sum of its matched sites)
    per_rid: dict[str, float] = {}
    for r in rows:
        if r["date"][:7] != as_of:
            continue
        rid = site_to_rid[r["refinery"]]
        if rid:
            per_rid[rid] = per_rid.get(rid, 0.0) + float(r["value"])

    stamped = 0
    for row in registry:
        rid = row["refinery_id"]
        if rid in per_rid:
            row["crude_capacity_kbd"] = f"{per_rid[rid]:g}"
            row["status"] = "operating" if per_rid[rid] > 0 else "shut"
            row["notes"] = f"capacity: EA nameplate {as_of}"
            stamped += 1
        if rid in COMPONENT_NOTES:
            row["notes"] = COMPONENT_NOTES[rid]

    reg_fields = ["refinery_id", "name", "owner", "city", "state", "padd",
                  "region", "crude_capacity_kbd", "status", "notes"]
    with registry_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=reg_fields)
        w.writeheader()
        for row in registry:
            w.writerow({k: row.get(k, "") for k in reg_fields})

    unmatched = {s: v for s, v in (
        (s, sum(float(r["value"]) for r in rows
                if r["refinery"] == s and r["date"][:7] == as_of))
        for s in sites if site_to_rid[s] is None
    ) if v > 0}
    unstamped = [row["refinery_id"] for row in registry
                 if row["refinery_id"] not in per_rid
                 and row["refinery_id"] not in COMPONENT_NOTES]
    return {
        "sites": len(sites),
        "matched_sites": sum(1 for v in site_to_rid.values() if v),
        "refineries_stamped": stamped,
        "us_total_kbd": round(sum(per_rid.values()), 1),
        "unmatched_active_sites": unmatched,
        "registry_without_capacity": unstamped,
        "monthly_csv": str(monthly_path),
        "as_of": as_of,
    }


# ------------------------------------------------------------- REM unit data
#
# REM export: one row per (unit type, refinery), capacity b/d by year
# 2010-2026. Refinery names are "City Operator" style; the alias map handles
# what automatic matching can't. None = deliberately skipped:
#   - combined entities REM merges but the registry splits (PBF East Coast =
#     Delaware City + Paulsboro; Marathon LA = Carson + Wilmington) - unit
#     stacks can't be split honestly, so those stay yield-mode;
#   - plants not in the registry (splitters, lube/asphalt, micro sites).

REM_ALIASES: dict[str, str | None] = {
    "Bayway": "PHILLIPS_LINDEN",
    "Big Spring": "ALON_BIG_SPRINGS",
    "Cherry Point": "BP_FERNDALE",
    "Commerce City": "SUNCOR_DENVER",
    "Corpus Christi Magellan": None,          # splitter, not in registry
    "East Coast Refining System PBF (Delaware and Paulsboro)": None,
    "El Dorado Frontier": "HF_EL_DORADO_KS",
    "El Dorado Lion Oil": "DELEK_EL_DORADO_AR",
    "Ely": "FORELAND_TONOPAH",
    "Evansville": "HF_CASPER",
    "Galena Park": "KINDER_HOUSTON",
    "Galveston Bay": "MPC_GALV_BAY",
    "Houston Petromax": None,                 # not in registry
    "Houston Targa Condensate Splitter": None,
    "Kapolei Par": "PAR_EWA_BEACH",
    "Kenai": "MARATHON_NIKISKI",
    "Kuparuk": "CONOCOPHILLIPS_ANCHORAGE",
    "Lake Charles Lubricants": None,          # lube plant
    "Lake Charles P66": "PHILLIPS_WESTLAKE",
    "Los Angeles Marathon": None,             # REM combines Carson+Wilmington
    "McKee": "VALERO_SUNRAY",
    "Navajo (Artesia)": "HF_ARTESIA",
    "Pine Bend": "FLINT_ROSEMOUNT",
    "Port Arthur Total": "TOTAL_PORT_ARTHUR",
    "Puget Sound": "HF_ANACORTES",
    "Salt Lake City BigWest": "BIG_NORTH_SLATERVILLE_CANAL",
    "Santa Maria": None,
    "Saraland": "VERTEX_MOBILE",
    "St. Charles": "VALERO_SAINT_CHARLES",
    "St. Paul Park": "MARATHON_SAINT_PAUL_PARK",
    "Wilmington Valero": "VALERO_LOS_ANGELES",
    "Wilmington Valero Asphalt": None,
    "Woods Cross Holly Corp": "HF_WOODS_CROSS",
}

# REM unit code -> (model unit_id, model unit_type). Codes not listed are not
# naphtha-relevant (HDS trains, BTX, lubes, H2, sulphur, ...) and are skipped.
REM_UNIT_MAP: dict[str, tuple[str, str]] = {
    "CDU": ("CDU", "CDU"),
    "VDU": ("VDU", "VDU"),
    "CCU/FCC": ("FCC", "FCC"),
    "RCC": ("RCC", "FCC"),
    "COK": ("COKER", "COKER"),
    "FCOK": ("FCOKER", "COKER"),
    "DHCU": ("DHCU", "HYDROCRACKER"),
    "RHCU": ("RHCU", "HYDROCRACKER"),
    "REF": ("REF", "REFORMER"),
    "CCR": ("CCR", "REFORMER"),
    "ISM": ("ISOM", "ISOM"),
    "ALK": ("ALKY", "ALKY"),
    "Naphtha Splitter": ("NSPL", "SPLITTER"),
    "HDS - NAP": ("NHT", "NHT"),
}

# RefineryDataTool process-unit name -> model unit_id (for utilization)
RDT_UNIT_MAP: dict[str, str] = {
    "CDU": "CDU", "VDU": "VDU", "FCC": "FCC", "RFCC": "RCC",
    "COKER": "COKER", "FCOKER": "FCOKER",
    "DHCU": "DHCU", "MHCU": "DHCU", "RHCU": "RHCU",
    "REF": "REF", "CCR": "CCR", "C5C6Isom": "ISOM", "ALKY": "ALKY",
    "NHT": "NHT", "Naphtha HT": "NHT",
}


def match_rem(name: str, operator: str, registry: list[dict]) -> str | None:
    """Resolve a REM refinery name ('City Operator' style) to a registry id."""
    if name in REM_ALIASES:
        return REM_ALIASES[name]
    n = _norm(name)
    cands = [r for r in registry if n.startswith(_norm(r["city"])) or _norm(r["city"]) == n]
    if len(cands) > 1:
        for r in cands:
            rest = n[len(_norm(r["city"])):].strip()
            if rest and rest in _norm(r["name"] + " " + r["owner"]):
                return r["refinery_id"]
        opn = _norm(operator)
        hinted = [r for r in cands
                  if opn and opn.split()[0] in _norm(r["name"] + " " + r["owner"])]
        if len(hinted) == 1:
            return hinted[0]["refinery_id"]
        return None
    return cands[0]["refinery_id"] if cands else None


def ingest_units(
    rem_csv: Path,
    rdt_csv: Path | None = None,
    data_dir: Path = DATA_DIR,
    year: str = "2026",
    util_year: str = "2024",
) -> dict:
    """Rebuild data/reference/units.csv from the REM unit-capacity export and
    (optionally) data/reference/unit_utilization.csv from the
    RefineryDataTool throughput export."""
    registry = _read_registry(data_dir / "reference" / "refineries.csv")

    rows = _read_utf16_tsv(rem_csv)
    ycol = rows[1].index(year)
    name_to_rid: dict[str, str | None] = {}
    units: dict[str, list[dict]] = {}
    skipped_active: set[str] = set()
    for r in rows[2:]:
        if len(r) <= ycol or not r[0] or not r[1]:
            continue
        code, name, operator = r[0], r[1], r[2]
        cap_kbd = _num(r[ycol]) / 1000.0
        if cap_kbd <= 0:
            continue
        if name not in name_to_rid:
            name_to_rid[name] = match_rem(name, operator, registry)
        rid = name_to_rid[name]
        if rid is None:
            if code == "CDU":
                skipped_active.add(name)
            continue
        if code not in REM_UNIT_MAP:
            continue
        unit_id, unit_type = REM_UNIT_MAP[code]
        units.setdefault(rid, []).append({
            "refinery_id": rid, "unit_id": unit_id, "unit_type": unit_type,
            "capacity_kbd": round(cap_kbd, 2),
            "notes": f"REM {year} capacity (source code: {code})",
        })

    units_path = data_dir / "reference" / "units.csv"
    with units_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["refinery_id", "unit_id", "unit_type",
                                           "capacity_kbd", "notes"])
        w.writeheader()
        for rid in sorted(units):
            w.writerows(units[rid])

    result = {
        "rem_refineries_matched": len({r for r in name_to_rid.values() if r}),
        "refineries_with_units": len(units),
        "unit_rows": sum(len(v) for v in units.values()),
        "skipped_active_rem_names": sorted(skipped_active),
        "units_csv": str(units_path),
    }

    if rdt_csv:
        rdt = _read_utf16_tsv(rdt_csv)
        hdr = rdt[0]
        idx = {h: hdr.index(h) for h in hdr}
        history: dict[tuple[str, str], dict[str, list[dict]]] = {}
        for r in rdt[1:]:
            if not r or not r[idx["Period"]]:
                continue
            unit_id = RDT_UNIT_MAP.get(r[idx["Process Unit"]])
            if unit_id is None:
                continue
            name = r[idx["Refinery Name"]]
            if name not in name_to_rid:
                name_to_rid[name] = match_rem(name, "", registry)
            rid = name_to_rid[name]
            if rid is None or not any(u["unit_id"] == unit_id for u in units.get(rid, [])):
                continue
            history.setdefault((rid, unit_id), {}).setdefault(r[idx["Period"]], []).append({
                "throughput_kbd": _num(r[idx["Unit Throughput bpd"]]) / 1000.0,
                "capacity_kbd": _num(r[idx["Unit Capacity bpd"]]) / 1000.0,
                "utilization": _num(r[idx["Unit Utilization Percent"]]) / 100.0,
            })

        def _avg(recs, key):
            return sum(x[key] for x in recs) / len(recs)

        # current-year actual utilization -> unit-level overrides
        util_path = data_dir / "reference" / "unit_utilization.csv"
        n_util = 0
        with util_path.open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["refinery_id", "unit_id", "year", "throughput_kbd",
                        "capacity_kbd", "utilization"])
            for (rid, unit_id), by_year in sorted(history.items()):
                recs = by_year.get(util_year)
                if not recs:
                    continue
                w.writerow([
                    rid, unit_id, util_year,
                    round(_avg(recs, "throughput_kbd"), 2),
                    round(_avg(recs, "capacity_kbd"), 2),
                    round(min(max(_avg(recs, "utilization"), 0.0), 1.1), 4),
                ])
                n_util += 1

        # effective capacity = demonstrated max annual throughput, excluding
        # the 2020 COVID year (nameplate is what's stated; this is what the
        # unit has actually proven it can run)
        nameplate = {(u["refinery_id"], u["unit_id"]): u["capacity_kbd"]
                     for us in units.values() for u in us}
        eff_path = data_dir / "reference" / "effective_capacity.csv"
        n_eff = 0
        with eff_path.open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["refinery_id", "unit_id", "nameplate_kbd",
                        "effective_kbd", "effective_year", "eff_vs_nameplate"])
            for (rid, unit_id), by_year in sorted(history.items()):
                candidates = [
                    (round(_avg(recs, "throughput_kbd"), 2), yr)
                    for yr, recs in by_year.items() if yr != "2020"
                ]
                if not candidates:
                    continue
                eff, yr = max(candidates)
                plate = nameplate.get((rid, unit_id), 0.0)
                w.writerow([rid, unit_id, plate, eff, yr,
                            round(eff / plate, 4) if plate else ""])
                n_eff += 1
        result["utilization_rows"] = n_util
        result["utilization_csv"] = str(util_path)
        result["effective_capacity_rows"] = n_eff
        result["effective_capacity_csv"] = str(eff_path)
    return result


# -------------------------------------------------------- reference extracts


def ingest_reference(data_dir: Path = DATA_DIR) -> dict:
    """Extract reference datasets from the raw REM / EA files:
    crude slate, 2021 product yields, US monthly naphtha balance."""
    registry = _read_registry(data_dir / "reference" / "refineries.csv")
    out = {}

    # crude slate (long format, non-empty cells only)
    slate = _read_utf16_tsv(data_dir / "raw" / "rem_crude_slate.csv")
    years = slate[1][4:]
    path = data_dir / "reference" / "crude_slate.csv"
    n = 0
    cache: dict[str, str | None] = {}
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["refinery_id", "rem_name", "crude_stream", "source_country",
                    "year", "slate_pct"])
        for r in slate[2:]:
            if not r or not r[0]:
                continue
            if r[0] not in cache:
                cache[r[0]] = match_rem(r[0], r[1], registry)
            for i, yr in enumerate(years):
                if len(r) > 4 + i and r[4 + i].strip():
                    w.writerow([cache[r[0]] or "", r[0], r[2], r[3], yr, r[4 + i]])
                    n += 1
    out["crude_slate_rows"] = n

    # 2021 full product yields
    y21 = _read_utf16_tsv(data_dir / "raw" / "rem_product_yields_2021.csv")
    path = data_dir / "reference" / "refinery_yields_2021.csv"
    n = 0
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["refinery_id", "rem_name", "product", "grade", "sub_region",
                    "yield_pct_2021"])
        for r in y21[2:]:
            if not r or not r[0] or not r[2]:
                continue
            if r[2] not in cache:
                cache[r[2]] = match_rem(r[2], "", registry)
            w.writerow([cache[r[2]] or "", r[2], r[0], r[1], r[3], r[4]])
            n += 1
    out["yields_2021_rows"] = n

    # feedstock slate with API / sulfur typing (RefineryDataTool1)
    fs = _read_utf16_tsv(data_dir / "raw" / "refinery_feedstock_slate_2023_2024.csv")
    hdr = fs[0]
    idx = {h: hdr.index(h) for h in hdr}
    path = data_dir / "reference" / "feedstock_slate.csv"
    n = 0
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["refinery_id", "rem_name", "year", "feedstock", "to_unit",
                    "kbd", "vol_pct", "api", "api_type", "sulfur_wt_pct",
                    "sulfur_type"])
        for r in fs[1:]:
            if not r or not r[idx["Period"]]:
                continue
            name = r[idx["Refinery Name"]]
            if name not in cache:
                cache[name] = match_rem(name, "", registry)
            w.writerow([
                cache[name] or "", name, r[idx["Period"]], r[idx["Feedstock"]],
                r[idx["To Unit"]],
                round(_num(r[idx["Barrel per Day"]]) / 1000.0, 2),
                round(_num(r[idx["Volume Percent"]]), 3),
                r[idx["API"]], r[idx["API Type"]],
                r[idx["Sulfur, wt%"]], r[idx["Sulfur Type"]],
            ])
            n += 1
    out["feedstock_rows"] = n

    # US monthly naphtha balance (EA workbook, Country_US tab)
    wb = openpyxl.load_workbook(
        data_dir / "raw" / "ea_naphtha_balance_2023_present.xlsx", data_only=True
    )
    ws = wb["Country_US"]
    path = data_dir / "reference" / "us_naphtha_balance_monthly.csv"
    n = 0
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["month", "flow", "kbd", "source"])
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] != "US":
                continue
            w.writerow([str(row[3])[:7], row[4], row[7], row[8]])
            n += 1
    out["us_balance_rows"] = n
    return out


# ---------------------------------------------------- offline events (IIR-style)
#
# The desk's offline-events export (OEVs sheet): planned & unplanned outage
# events per unit with dates and offline capacity. Ingestion writes
#   data/reference/outage_events.csv    all US events ending 2023+ (matched)
#   data/reference/current_outages.csv  offline fraction per refinery/unit
#                                       as of a given date (feeds the boxes'
#                                       OFFLINE % prefill)

OEV_UNIT_MAP: dict[str, str] = {
    "Atmospheric Distillation": "CDU",
    "Vacuum Distillation": "VDU",
    "FCCU (Fluid Catalytic Cracker)": "FCC",
    "Delayed Coker": "COKER",
    "Fluid Coker": "FCOKER",
    "Flexicoker": "FCOKER",
    "Distillate Hydrocracker": "DHCU",
    "Resid Hydrocracker": "RHCU",
    "Semiregen/Cyclic Reformer": "REF",
    "CCR (Continuous Catalytic Reformer)": "CCR",
    "Isomerization": "ISOM",
    "Hydrofluoric Alkylation": "ALKY",
    "Sulfuric Alkylation": "ALKY",
    "Reformer Feed Hydrotreater": "NHT",
    "Light Naphtha Hydrotreater": "NHT",
    "Naphtha Hydrotreater": "NHT",
}

# plant-name|city -> refinery_id; None = not modelled (splitters not yet in
# the registry, lube/wax/asphalt plants, shut sites). Cities often differ
# between databases (Joliet=Channahon, Pine Bend=Inver Grove Heights, ...).
EXXON_PLANT_ALIASES: dict[str, str | None] = {
    "Big Spring Refinery|Big Spring": "ALON_BIG_SPRINGS",
    "Channelview Condensate Splitter|Channelview": None,
    "Cherry Point Refinery|Blaine": "BP_FERNDALE",
    "Corpus Christi Condensate Splitter (ONEOK)|Corpus Christi": None,
    "Denver Refinery - Plant 1 (West)|Commerce City": "SUNCOR_DENVER",
    "Denver Refinery - Plant 2 (East)|Commerce City": "SUNCOR_DENVER",
    "Eagle Springs Refinery|Ely": "FORELAND_TONOPAH",
    "El Dorado Refinery (Delek)|El Dorado": "DELEK_EL_DORADO_AR",
    "El Dorado Refinery (HF Sinclair)|El Dorado": "HF_EL_DORADO_KS",
    "Galena Park Natural Gas Condensate Splitter|Galena Park": "KINDER_HOUSTON",
    "Galveston Crude Refinery|Galveston": None,
    "Houston Fractionation Refinery|Houston": None,
    "Houston Refinery (Houston Refining)|Houston": "LYONDELLBASELL_HOUSTON",
    "Joliet Refinery|Channahon": "EXXONMOBIL_JOLIET",
    "Kenai Refinery|Kenai": "MARATHON_NIKISKI",
    "Los Angeles Refinery - Wilmington (Marathon)|Wilmington": "MARATHON_LOS_ANGELES",
    "Los Angeles Refinery - Wilmington|Wilmington": "PHILLIPS_LOS_ANGELES",
    "Lovington Refinery|Lovington": "HF_ARTESIA",   # Artesia/Lovington complex
    "Marrero Re-Refined Oil Lubricants|Marrero": None,
    "Mobile Refinery (Vertex)|Saraland": "VERTEX_MOBILE",
    "North Salt Lake Refinery|North Salt Lake": "BIG_NORTH_SLATERVILLE_CANAL",
    "Par East Refinery|Kapolei": "PAR_EWA_BEACH",
    "Pine Bend Refinery|Inver Grove Heights": "FLINT_ROSEMOUNT",
    "Rodeo Refinery|Rodeo": None,                   # renewables conversion
    "Salt Lake City Refinery (Chevron)|North Salt Lake": "CHEVRON_SALT_LAKE_CITY",
    "Sandersville Refinery|Heidelberg": "HUNT_SANDERSVILLE",
    "Smethport Waxes|Smethport": None,
    "Sweeny Refinery|Old Ocean": "PHILLIPS_SWEENY",
    "Toledo Refinery (Cenovus)|Oregon": "CENOVUS_TOLEDO",
    "Toledo Refinery (PBF)|Oregon": "PBF_TOLEDO",
    "Wilmington Asphalt Refinery|Wilmington": None,
    "Wilmington Refinery|Wilmington": "VALERO_LOS_ANGELES",
    "Wood River Refinery|Roxana": "CENOVUS_WOOD_RIVER",
}


def _match_plant(name: str, city: str, state: str, owner: str,
                 registry: list[dict]) -> str | None:
    key = f"{name}|{city}"
    if key in EXXON_PLANT_ALIASES:
        return EXXON_PLANT_ALIASES[key]
    n_city = _norm(city or "")
    cands = [r for r in registry if _norm(r["city"]) == n_city]
    if len(cands) > 1:
        hints = _norm(f"{name} {owner}")
        hinted = [r for r in cands
                  if any(tok in hints for tok in _norm(r["name"]).split()
                         if len(tok) > 3)]
        if len(hinted) == 1:
            return hinted[0]["refinery_id"]
        strong = [r for r in cands
                  if _norm(r["owner"]).split()[0] in hints
                  or _norm(r["name"]).split()[0] in hints]
        if len(strong) == 1:
            return strong[0]["refinery_id"]
        return None
    return cands[0]["refinery_id"] if cands else None


def ingest_outages(
    xlsx_path: Path, data_dir: Path = DATA_DIR, as_of: str = "2026-07-09"
) -> dict:
    """OEVs from the desk's IIR export workbook -> outage CSVs."""
    wb = openpyxl.load_workbook(Path(xlsx_path), data_only=True, read_only=True)
    ws = wb["OEVs"]
    rows = ws.iter_rows(min_row=1, values_only=True)
    hdr = [str(h) for h in next(rows)]
    records = [dict(zip(hdr, row)) for row in rows]
    return ingest_outage_records(records, data_dir=data_dir, as_of=as_of)


def ingest_outage_records(
    records: list[dict], data_dir: Path = DATA_DIR, as_of: str = "2026-07-09"
) -> dict:
    """IIR OEV records (flattened dicts, xlsx export or live API pull) ->
    outage_events.csv + current_outages.csv."""
    registry = _read_registry(data_dir / "reference" / "refineries.csv")

    def d(v):  # '2023-03-25T05:00:00Z[UTC]' -> '2023-03-25'
        return str(v)[:10] if v else ""

    plant_cache: dict = {}
    events = []
    unmatched_plants = set()
    for rec in records:
        if rec.get("plantPhysicalAddress.countryName") not in (
                "U.S.A.", "United States"):
            continue
        end = d(rec.get("eventEndDate"))
        if end < "2023-01-01":
            continue
        pid = rec.get("associatedPlantId")
        if pid not in plant_cache:
            plant_cache[pid] = _match_plant(
                str(rec.get("plantName") or ""),
                str(rec.get("plantPhysicalAddress.city") or ""),
                str(rec.get("plantPhysicalAddress.stateName") or ""),
                str(rec.get("plantOwnerName") or ""), registry)
        rid = plant_cache[pid]
        if rid is None:
            unmatched_plants.add(
                f'{rec.get("plantName")} ({rec.get("plantPhysicalAddress.city")}, '
                f'{rec.get("plantPhysicalAddress.stateName")})')
        events.append({
            "refinery_id": rid or "",
            "event_id": rec.get("eventId"),
            "event_type": rec.get("eventType"),
            "status": rec.get("derivedEventStatusDesc"),
            "start": d(rec.get("eventStartDate")),
            "end": end,
            "duration_days": rec.get("eventDuration"),
            "unit_name": rec.get("unitName"),
            "unit_type": rec.get("unitTypeDesc"),
            "model_unit": OEV_UNIT_MAP.get(str(rec.get("unitTypeDesc")), ""),
            "unit_capacity": rec.get("offlineCapacity.unitCapacity"),
            "capacity_offline": rec.get("offlineCapacity.capacityOffline"),
            "uom": rec.get("offlineCapacity.uom"),
            "confirmation": rec.get("eventConfirmationStatus"),
            "cause": rec.get("eventCause"),
            "plant_name": rec.get("plantName"),
            "city": rec.get("plantPhysicalAddress.city"),
            "state": rec.get("plantPhysicalAddress.stateName"),
        })

    ev_path = data_dir / "reference" / "outage_events.csv"
    fields = list(events[0].keys())
    with ev_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        w.writerows(sorted(events, key=lambda e: (e["start"], str(e["refinery_id"]))))

    # as-of snapshot: offline fraction per (refinery, model unit)
    groups: dict[tuple[str, str], list] = {}
    for e in events:
        if not e["refinery_id"] or not e["model_unit"]:
            continue
        if not (e["start"] <= as_of <= e["end"]):
            continue
        cap = float(e["unit_capacity"] or 0)
        off = float(e["capacity_offline"] or 0)
        if cap > 0:
            groups.setdefault((e["refinery_id"], e["model_unit"]), []).append(
                (off, cap, e["event_type"], e["event_id"]))
    cur_path = data_dir / "reference" / "current_outages.csv"
    with cur_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["refinery_id", "unit_id", "as_of", "offline_frac",
                    "events", "event_types"])
        for (rid, unit), recs in sorted(groups.items()):
            frac = min(1.0, sum(o for o, *_ in recs) / max(sum(c for _, c, *_ in recs), 1e-9))
            w.writerow([rid, unit, as_of, round(frac, 4),
                        ";".join(str(x[3]) for x in recs),
                        ";".join(sorted({x[2] for x in recs}))])

    matched_plants = sum(1 for v in plant_cache.values() if v)
    return {
        "us_events_2023_plus": len(events),
        "plants_matched": f"{matched_plants}/{len(plant_cache)}",
        "current_outage_rows": len(groups),
        "unmatched_plants": sorted(unmatched_plants),
        "events_csv": str(ev_path),
        "current_csv": str(cur_path),
        "as_of": as_of,
    }
