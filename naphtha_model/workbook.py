"""Desk workbook builder.

Generates a live, formula-driven Excel model: traders edit blue/orange input
cells (assumptions, outages, flows, scenario toggle) directly in Excel and
every downstream number — refinery boxes, PADD balance, charts, checks —
recalculates without touching Python. Python is the *generator/refresher*:
it seeds the workbook from the data/ directory and wires the formulas.

Tabs
----
README       how to drive the model, color legend
Model        settings: forward start date, scenario toggle
Dashboard    line/bar charts: supply forecast vs demand, balance, outage
             at-risk, cargo flows, per-refinery forecast
Boxes        one whiteboard-style box per refinery (units, capacity,
             utilization, yields, weekly forward strip) — all live formulas
Balance      per-PADD weekly supply/flows/demand/balance/at-risk
Assumptions  per-PADD utilization + yield matrices (inputs)
Outages      planned / unplanned / scenario rows (inputs)
Flows        ship-tracking / cargo table (inputs)
Demand       demand by PADD & sector (inputs)
Refineries   registry (reference)
Intel        market intel log (reference)
Checks       automated data-quality checks with PASS/FAIL status
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

from .config import DATA_DIR, PADDS, UNIT_TYPES
from .loaders import ModelData
from .schema import ProcessUnit, Refinery

# ------------------------------------------------------------------ styling
#
# House theme: deep navy + gold accent, white-on-steel headers, no
# gridlines, color-coded tabs. Inputs stay Excel-convention blue, manual
# overrides orange, formulas soft grey.

NAVY = "16243D"           # banners / brand
STEEL = "27406B"          # table headers
GOLD = "C9A227"           # accent
FILL_BANNER = PatternFill("solid", fgColor=NAVY)
FILL_HDR = PatternFill("solid", fgColor=STEEL)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
FILL_INPUT = PatternFill("solid", fgColor="DCE9F7")     # blue: editable input
FILL_OVERRIDE = PatternFill("solid", fgColor="FCE4D6")  # orange: manual override
FILL_CALC = PatternFill("solid", fgColor="F4F6FA")      # grey: formula, don't type
FILL_TOTAL = PatternFill("solid", fgColor="F3E8C8")     # soft gold: totals
FILL_PASS = PatternFill("solid", fgColor="C6EFCE")
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")

FONT_BANNER = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
FONT_HDR = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
FONT_SMALL = Font(size=9, name="Calibri")
THIN = Side(style="thin", color="D5DAE3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD_EDGE = Side(style="medium", color=GOLD)

TAB_COLORS = {
    "Cover": GOLD,
    "Data": "4472C4", "Assumptions": "4472C4",          # inputs: blue
    "Boxes": NAVY, "Individual Refineries": NAVY,                                       # the engine
    "Nameplate": "2E8677", "Effective": "2E8677",        # capacity views: teal
    "Outages": "B03A2E", "Outage History": "7A2A22",
    "Kpler Flows": "1F6FB2", "US Balance": "216E5E",
    "Live Feeds": "B27C1F",                              # data in: amber
    "CrudeSlate": "6B8E23",                              # slate: olive
    "BlendEcon": "A23B3B",                               # economics: maroon
    "KitWalk": "7A5195",                                 # tuning: plum
}


def _theme(ws, tab_color: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    if tab_color:
        ws.sheet_properties.tabColor = tab_color

# ------------------------------------------------------- scan-range budgets
# Formulas scan fixed ranges so traders can append rows without re-wiring.

OUT_LO, OUT_HI = 2, 201        # Outages data rows
FLOW_LO, FLOW_HI = 2, 300      # Flows data rows
DEM_LO, DEM_HI = 2, 100        # Demand data rows
REF_LO, REF_HI = 2, 300        # Refineries registry rows
BOX_LO, BOX_HI = 4, 2600       # Boxes data rows




def _axes(ch, x_title: str, y_title: str, rotate: bool = False,
          skip: int | None = None) -> None:
    """Readable charts: visible titled axes, no gridlines, spaced bars,
    rotated/skipped date labels on dense axes."""
    from openpyxl.chart import BarChart
    from openpyxl.drawing.text import (CharacterProperties, Paragraph,
                                       ParagraphProperties, RichTextProperties)
    from openpyxl.chart.text import RichText

    ch.x_axis.title = x_title
    ch.y_axis.title = y_title
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.x_axis.tickLblPos = "low"
    ch.y_axis.tickLblPos = "nextTo"
    ch.y_axis.majorGridlines = None          # kill the grid clutter
    if isinstance(ch, BarChart):
        ch.gapWidth = 60                     # fatter bars
        if len(ch.series) > 1:
            ch.overlap = -5
    if rotate:
        ch.x_axis.txPr = RichText(
            bodyPr=RichTextProperties(rot=-2700000, vert="horz"),
            p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=800)),
                         endParaRPr=CharacterProperties())],
        )
    if skip:
        ch.x_axis.tickLblSkip = skip
        ch.x_axis.tickMarkSkip = skip
    # thicken line series and make sure every bar series has a fill
    for ser in ch.series:
        gp = ser.graphicalProperties
        if not isinstance(ch, BarChart):
            if gp.line is not None:
                gp.line.width = 26000        # ~2pt
        elif gp.solidFill is None:
            gp.solidFill = NAVY

def _banner(ws, row: int, text: str, span: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 20
    c = ws.cell(row=row, column=1, value=text)
    c.fill = FILL_BANNER
    c.font = FONT_BANNER
    c.alignment = Alignment(vertical="center")
    edge = Border(bottom=GOLD_EDGE)
    for col in range(1, span + 1):
        cc = ws.cell(row=row, column=col)
        cc.fill = FILL_BANNER
        cc.border = edge


def _hdr(ws, row: int, col: int, text: str) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = FONT_HDR
    c.fill = FILL_HDR
    c.border = BORDER
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _style(cell, fill=None, fmt=None, bold=False, border=True):
    if fill is not None:
        cell.fill = fill
    if fmt is not None:
        cell.number_format = fmt
    if bold:
        cell.font = Font(bold=True, size=9, name="Calibri")
    else:
        cell.font = FONT_SMALL
    if border:
        cell.border = BORDER
    return cell


class DeskWorkbook:
    """Builds the workbook. One instance per export."""

    def __init__(self, data: ModelData, axis: list[date], scenarios=None):
        self.scenarios = scenarios or []
        self.data = data
        self.book = data.book
        self.axis = axis
        self.weeks = len(axis)
        self.cuts = list(self.book.cuts)
        self.padds = sorted({r.padd for r in data.refineries if r.region == "US"})
        self.wb = Workbook()

        # Boxes column map (1-indexed), dynamic in the number of cuts:
        # A type | B refinery | C padd | D unit | E unit_type | F cap |
        # G eff cap | H offline% | I mode toggle | J util_ovr | K util |
        # (ovr, yield) per cut | ysum | base | net wk1..wkW | off% wk1..wkW
        self.c_type, self.c_rid, self.c_padd, self.c_uid, self.c_utype = 1, 2, 3, 4, 5
        self.c_cap, self.c_effcap, self.c_out = 6, 7, 8
        self.c_mode, self.c_utilov, self.c_util = 9, 10, 11
        self.c_cut0 = 12                               # first cut override col
        self.c_ysum = self.c_cut0 + 2 * len(self.cuts)
        self.c_base = self.c_ysum + 1
        self.c_net0 = self.c_base + 1                  # first weekly net col
        self.c_off0 = self.c_net0 + self.weeks         # first weekly off% col
        self.c_last = self.c_off0 + self.weeks - 1

        # effective capacity (demonstrated max annual throughput) if ingested
        self.eff_caps: dict[tuple[str, str], float] = {}
        self.eff_years: dict[tuple[str, str], str] = {}
        eff_path = DATA_DIR / "reference" / "effective_capacity.csv"
        if eff_path.exists():
            import csv as _csv

            with eff_path.open() as fh:
                for row in _csv.DictReader(fh):
                    key = (row["refinery_id"], row["unit_id"])
                    self.eff_caps[key] = float(row["effective_kbd"])
                    self.eff_years[key] = row["effective_year"]

        # populated as sheets are built
        self.assump_refs: dict = {}
        self.balance_rows: dict[int, dict[str, int]] = {}
        self.balance_week_row: dict[int, int] = {}
        self.refinery_total_rows: dict[str, int] = {}

    # -------------------------------------------------------------- helpers

    def _cut_ovr_col(self, i: int) -> int:
        return self.c_cut0 + 2 * i

    def _cut_yld_col(self, i: int) -> int:
        return self.c_cut0 + 2 * i + 1

    def _resolved_padd(self, padd: int):
        """Resolved (util, {unit_type: {cut: yield}}) for a PADD on day one,
        ignoring refinery-specific overrides (dummy refinery)."""
        dummy = Refinery(
            refinery_id="__PADD__", name="", owner="", city="", state="",
            padd=padd, crude_capacity_kbd=0,
        )
        day = self.axis[0]
        yields = {}
        util = None
        for ut in UNIT_TYPES:
            unit = ProcessUnit("__PADD__", "__U__", ut, 0.0)
            if util is None:
                util = self.book.utilization(dummy, unit, day).value
            yields[ut] = {
                cut: self.book.yield_for(dummy, unit, cut, day).value
                for cut in self.cuts
            }
        return util, yields

    # ---------------------------------------------------------------- build

    def build(self, out_path: Path) -> Path:
        self._sheet_model()
        self._sheet_assumptions()
        self._sheet_refineries()
        self._sheet_outages()
        self._sheet_flows()
        self._sheet_demand()
        self._sheet_intel()
        self._sheet_boxes()
        self._sheet_balance()
        self._sheet_calibration()
        self._sheet_yields_2024()
        self._sheet_dashboard()
        self._sheet_checks()
        self._sheet_readme()

        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        order = [
            "README", "Dashboard", "Boxes", "Balance", "Calibration",
            "Assumptions", "Outages", "Flows", "Demand", "Refineries",
            "Yields_2024", "Intel", "Checks", "Model",
        ]
        self.wb._sheets = [self.wb[name] for name in order]
        out_path = Path(out_path)
        self.wb.save(out_path)
        return out_path

    # ---------------------------------------------------------------- Model

    def _sheet_model(self) -> None:
        ws = self.wb.create_sheet("Model")
        _banner(ws, 1, "Model settings", 4)
        rows = [
            ("Forward window start (edit to shift every date in the book)", self.axis[0], "yyyy-mm-dd", FILL_INPUT),
            ("Weeks in window (fixed at generation; regenerate to change)", self.weeks, "0", FILL_CALC),
            ("Include SCENARIO outage rows? (YES/NO)", "NO", None, FILL_INPUT),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M"), None, FILL_CALC),
            ("Regenerate with", "python -m naphtha_model export --out <file>", None, FILL_CALC),
        ]
        for i, (label, value, fmt, fill) in enumerate(rows, start=3):
            ws.cell(row=i, column=1, value=label).font = FONT_SMALL
            _style(ws.cell(row=i, column=2, value=value), fill=fill, fmt=fmt)
        dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add("B5")
        ws.column_dimensions["A"].width = 55
        ws.column_dimensions["B"].width = 40
        # cell used by formulas elsewhere
        self.start_ref = "Model!$B$3"
        self.toggle_ref = "Model!$B$5"

    # ---------------------------------------------------------- Assumptions

    def _sheet_assumptions(self) -> None:
        ws = self.wb.create_sheet("Assumptions")
        span = 6 + len(self.cuts)
        _banner(
            ws, 1,
            "Assumptions — blue cells are inputs. Yields are % of unit throughput; "
            "consumers (reformer/isom) are NEGATIVE. Edits flow straight into Boxes/Balance.",
            span + 8,
        )
        # long lookup tables (right-hand side) that Boxes formulas read
        UL_PADD, UL_VAL = span + 2, span + 3            # util long: padd, value
        YL_PADD, YL_UT, YL_CUT, YL_VAL = span + 5, span + 6, span + 7, span + 8
        for col, label in [
            (UL_PADD, "padd"), (UL_VAL, "util"),
            (YL_PADD, "padd"), (YL_UT, "unit_type"), (YL_CUT, "cut"), (YL_VAL, "yield"),
        ]:
            _hdr(ws, 3, col, label)

        ul_row, yl_row = 4, 4
        r = 3
        for padd in self.padds:
            util, yields = self._resolved_padd(padd)
            _banner(ws, r, f"PADD {padd} — {PADDS.get(padd, '')}", span)
            ws.cell(row=r + 1, column=1, value="Utilization").font = FONT_HDR
            ucell = ws.cell(row=r + 1, column=2, value=util)
            _style(ucell, fill=FILL_INPUT, fmt="0.0%")
            # util long row points at the matrix cell
            ws.cell(row=ul_row, column=UL_PADD, value=padd)
            ws.cell(row=ul_row, column=UL_VAL, value=f"={ucell.coordinate}")
            ul_row += 1

            _hdr(ws, r + 3, 1, "Unit type")
            for j, cut in enumerate(self.cuts):
                _hdr(ws, r + 3, 2 + j, cut)
            for i, ut in enumerate(UNIT_TYPES):
                row = r + 4 + i
                ws.cell(row=row, column=1, value=ut).font = FONT_SMALL
                for j, cut in enumerate(self.cuts):
                    cell = ws.cell(row=row, column=2 + j, value=yields[ut][cut])
                    _style(cell, fill=FILL_INPUT, fmt="0.0%")
                    ws.cell(row=yl_row, column=YL_PADD, value=padd)
                    ws.cell(row=yl_row, column=YL_UT, value=ut)
                    ws.cell(row=yl_row, column=YL_CUT, value=cut)
                    ws.cell(row=yl_row, column=YL_VAL, value=f"={cell.coordinate}")
                    yl_row += 1
            r += 5 + len(UNIT_TYPES) + 1

        ul_col_p, ul_col_v = get_column_letter(UL_PADD), get_column_letter(UL_VAL)
        yl_p, yl_ut = get_column_letter(YL_PADD), get_column_letter(YL_UT)
        yl_c, yl_v = get_column_letter(YL_CUT), get_column_letter(YL_VAL)
        self.assump_refs = {
            "util": (
                f"Assumptions!${ul_col_v}$4:${ul_col_v}${ul_row - 1}",
                f"Assumptions!${ul_col_p}$4:${ul_col_p}${ul_row - 1}",
            ),
            "yield": (
                f"Assumptions!${yl_v}$4:${yl_v}${yl_row - 1}",
                f"Assumptions!${yl_p}$4:${yl_p}${yl_row - 1}",
                f"Assumptions!${yl_ut}$4:${yl_ut}${yl_row - 1}",
                f"Assumptions!${yl_c}$4:${yl_c}${yl_row - 1}",
            ),
        }
        ws.column_dimensions["A"].width = 16
        ws.freeze_panes = "A2"

    # ------------------------------------------------------------ Refineries

    def _sheet_refineries(self) -> None:
        ws = self.wb.create_sheet("Refineries")
        _banner(ws, 1, "Refinery registry (reference — edit via data/reference and regenerate)", 9)
        headers = ["refinery_id", "name", "owner", "city", "state", "padd",
                   "crude_kbd", "status", "notes"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 2, c, h)
        for i, r in enumerate(self.data.refineries, start=3):
            vals = [r.refinery_id, r.name, r.owner, r.city, r.state, r.padd,
                    r.crude_capacity_kbd, r.status, r.notes]
            for c, v in enumerate(vals, start=1):
                _style(ws.cell(row=i, column=c, value=v))
        widths = [16, 24, 20, 14, 6, 6, 10, 10, 44]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

    # -------------------------------------------------------------- Outages

    def _sheet_outages(self) -> None:
        ws = self.wb.create_sheet("Outages")
        headers = ["category", "refinery_id", "unit_id", "outage_id", "start",
                   "end", "offline_pct", "active", "source", "confidence", "notes"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 1, c, h)
        ws.cell(row=1, column=12, value=(
            "Add a row the moment an outage hits — Boxes/Balance/Dashboard update "
            "instantly. category=scenario rows only count when Model!B5 = YES."
        )).font = FONT_SMALL

        scenario_outages = [
            (s.name, o) for s in self.scenarios for o in s.outages
        ]
        row = OUT_LO
        for o in self.data.outages:
            vals = [o.outage_type, o.refinery_id, o.unit_id, o.outage_id,
                    o.start, o.end, o.offline_pct, None, o.source, o.confidence, o.notes]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                fill = FILL_CALC if c == 8 else FILL_INPUT
                fmt = "yyyy-mm-dd" if c in (5, 6) else ("0" if c == 7 else None)
                _style(cell, fill=fill, fmt=fmt)
            row += 1
        for scn_name, o in scenario_outages:
            vals = ["scenario", o.refinery_id, o.unit_id, o.outage_id,
                    o.start, o.end, o.offline_pct, None, scn_name, "", o.notes]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                fill = FILL_CALC if c == 8 else FILL_INPUT
                fmt = "yyyy-mm-dd" if c in (5, 6) else ("0" if c == 7 else None)
                _style(cell, fill=fill, fmt=fmt)
            row += 1
        # blank input rows + the active formula down the whole scan range
        for rr in range(OUT_LO, OUT_HI + 1):
            cell = ws.cell(
                row=rr, column=8,
                value=f'=IF($B{rr}="","",IF($A{rr}="scenario",{self.toggle_ref},"YES"))',
            )
            _style(cell, fill=FILL_CALC)
            if rr >= row:
                for c in (1, 2, 3, 4, 5, 6, 7, 9, 10, 11):
                    fmt = "yyyy-mm-dd" if c in (5, 6) else None
                    _style(ws.cell(row=rr, column=c), fill=FILL_INPUT, fmt=fmt)

        dv = DataValidation(type="list", formula1='"planned,unplanned,scenario"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"A{OUT_LO}:A{OUT_HI}")
        widths = [11, 14, 9, 14, 11, 11, 11, 8, 16, 11, 40]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    # ---------------------------------------------------------------- Flows

    def _sheet_flows(self) -> None:
        ws = self.wb.create_sheet("Flows")
        headers = ["date", "padd", "direction", "volume_kbd", "cut",
                   "counterparty", "vessel", "source", "notes", "category", "signed_kbd"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 1, c, h)
        ws.cell(row=1, column=12, value=(
            "Ship-tracking / cargo table. import & transfer_in ADD to the PADD, "
            "export & transfer_out REMOVE. category=scenario rows only count when "
            "Model!B5 = YES. signed_kbd is automatic."
        )).font = FONT_SMALL

        rows = [(f, "") for f in self.data.flows] + [
            (f, "scenario") for s in self.scenarios for f in s.flows
        ]
        row = FLOW_LO
        for f, category in rows:
            vals = [f.flow_date, f.padd, f.direction, f.volume_kbd, f.cut,
                    f.counterparty, f.vessel, f.source, f.notes, category]
            for c, v in enumerate(vals, start=1):
                _style(ws.cell(row=row, column=c, value=v),
                       fill=FILL_INPUT, fmt="yyyy-mm-dd" if c == 1 else None)
            row += 1
        for rr in range(FLOW_LO, FLOW_HI + 1):
            cell = ws.cell(
                row=rr, column=11,
                value=(f'=IF($C{rr}="","",'
                       f'IF(AND($J{rr}="scenario",{self.toggle_ref}<>"YES"),0,'
                       f'IF(OR($C{rr}="import",$C{rr}="transfer_in"),$D{rr},-$D{rr})))'),
            )
            _style(cell, fill=FILL_CALC, fmt="0.0")
            if rr >= row:
                for c in range(1, 11):
                    _style(ws.cell(row=rr, column=c), fill=FILL_INPUT,
                           fmt="yyyy-mm-dd" if c == 1 else None)

        dv = DataValidation(type="list",
                            formula1='"import,export,transfer_in,transfer_out"',
                            allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"C{FLOW_LO}:C{FLOW_HI}")
        dv2 = DataValidation(type="list", formula1='"scenario"', allow_blank=True)
        ws.add_data_validation(dv2)
        dv2.add(f"J{FLOW_LO}:J{FLOW_HI}")
        widths = [11, 6, 12, 11, 8, 16, 16, 14, 30, 10, 11]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    # --------------------------------------------------------------- Demand

    def _sheet_demand(self) -> None:
        ws = self.wb.create_sheet("Demand")
        headers = ["padd", "sector", "volume_kbd", "start", "end", "notes"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 1, c, h)
        ws.cell(row=1, column=7, value=(
            "Steady-state naphtha dispositions per PADD. The workbook treats every "
            "row as always-on; dated demand is applied by the Python engine."
        )).font = FONT_SMALL
        for i, d in enumerate(self.data.demand, start=DEM_LO):
            vals = [d.padd, d.sector, d.volume_kbd, d.start, d.end, d.notes]
            for c, v in enumerate(vals, start=1):
                _style(ws.cell(row=i, column=c, value=v),
                       fill=FILL_INPUT, fmt="yyyy-mm-dd" if c in (4, 5) else None)
        widths = [6, 20, 11, 11, 11, 44]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

    # ---------------------------------------------------------------- Intel

    def _sheet_intel(self) -> None:
        ws = self.wb.create_sheet("Intel")
        headers = ["date", "refinery_id", "padd", "headline", "impact_kbd",
                   "linked_outage_id", "source", "confidence", "notes"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 1, c, h)
        for i, n in enumerate(self.data.intel, start=2):
            vals = [n.note_date, n.refinery_id, n.padd, n.headline, n.impact_kbd,
                    n.linked_outage_id, n.source, n.confidence, n.notes]
            for c, v in enumerate(vals, start=1):
                _style(ws.cell(row=i, column=c, value=v),
                       fill=FILL_INPUT, fmt="yyyy-mm-dd" if c == 1 else None)
        widths = [11, 14, 6, 60, 11, 15, 12, 11, 40]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

    # ---------------------------------------------------------------- Boxes

    def _offline_formula(self, row: int, week_idx: int, ws_name: str = "Boxes") -> str:
        """Prorated offline fraction for one unit-week: sum over matching active
        outage rows of offline% x overlap_days/7, capped at 1. Elementwise
        min/max built from ABS so it works inside SUMPRODUCT."""
        net_col = get_column_letter(self.c_net0 + week_idx)
        ws_ = f"${net_col}$2"                     # week start date (header)
        we = f"(${net_col}$2+6)"                  # week end date
        S = f"Outages!$E${OUT_LO}:$E${OUT_HI}"
        E = f"Outages!$F${OUT_LO}:$F${OUT_HI}"
        min_e_we = f"(({E}+{we}-ABS({E}-{we}))/2)"
        max_s_ws = f"(({S}+{ws_}+ABS({S}-{ws_}))/2)"
        od = f"({min_e_we}-{max_s_ws}+1)"
        clip = f"(({od}+ABS({od}))/2)"
        flags = (
            f"(Outages!$B${OUT_LO}:$B${OUT_HI}=$B{row})"
            f"*((Outages!$C${OUT_LO}:$C${OUT_HI}=$D{row})"
            f"+(Outages!$C${OUT_LO}:$C${OUT_HI}=\"\"))"
            f"*(Outages!$H${OUT_LO}:$H${OUT_HI}=\"YES\")"
        )
        return (
            f"=MIN(1,SUMPRODUCT({flags}"
            f"*(Outages!$G${OUT_LO}:$G${OUT_HI}/100)*{clip}/7))"
        )

    def _sheet_boxes(self) -> None:
        ws = self.wb.create_sheet("Boxes")
        _banner(
            ws, 1,
            "Refinery boxes — every refinery, unit by unit. Blue = capacity input, "
            "orange = manual overrides (clear the cell to fall back to PADD/global "
            "assumptions), grey = live formulas. Weekly strip nets outages automatically.",
            self.c_last,
        )
        # week date headers (row 2) over both the net strip and the off% strip
        for k in range(self.weeks):
            for c0 in (self.c_net0, self.c_off0):
                cell = ws.cell(row=2, column=c0 + k, value=f"={self.start_ref}+{k}*7")
                _style(cell, fill=FILL_CALC, fmt="m/d", bold=True)
        # static headers (row 3)
        labels = {
            self.c_type: "row", self.c_rid: "refinery_id", self.c_padd: "padd",
            self.c_uid: "unit", self.c_utype: "type", self.c_cap: "cap kbd",
            self.c_effcap: "eff cap", self.c_out: "offline %",
            self.c_mode: "mode", self.c_utilov: "util ovr", self.c_util: "util",
            self.c_ysum: "Σ yield", self.c_base: "base net kbd",
        }
        for i, cut in enumerate(self.cuts):
            labels[self._cut_ovr_col(i)] = f"{cut} ovr"
            labels[self._cut_yld_col(i)] = f"{cut} yield"
        for col, text in labels.items():
            _hdr(ws, 3, col, text)
        for k in range(self.weeks):
            _hdr(ws, 3, self.c_net0 + k, "net kbd")
            _hdr(ws, 3, self.c_off0 + k, "off%")

        util_val, util_padd = self.assump_refs["util"]
        y_val, y_padd, y_ut, y_cut = self.assump_refs["yield"]
        row = BOX_LO
        day1 = self.axis[0]

        for ref in sorted(self.data.refineries, key=lambda r: (r.padd, -r.crude_capacity_kbd)):
            _banner(
                ws, row,
                f"  {ref.name}  [{ref.refinery_id}]   —   {ref.owner}   —   PADD {ref.padd}"
                f"   —   crude {ref.crude_capacity_kbd:,.0f} kbd",
                self.c_last,
            )
            row += 1
            first_unit = row
            ovl = get_column_letter(self.c_utilov)
            mdl = get_column_letter(self.c_mode)
            outl = get_column_letter(self.c_out)
            total_row = first_unit + len(ref.units)
            offx = (f"(1-MIN(1,MAX(N(${outl}{{row}}),"
                    f"N(${outl}{total_row}))))")
            mode = f'IF(${mdl}{total_row}<>"",${mdl}{total_row},${mdl}{{row}})'
            for unit in ref.units:
                ws.cell(row=row, column=self.c_type, value="UNIT").font = FONT_SMALL
                _style(ws.cell(row=row, column=self.c_rid, value=ref.refinery_id))
                _style(ws.cell(row=row, column=self.c_padd, value=ref.padd))
                _style(ws.cell(row=row, column=self.c_uid, value=unit.unit_id))
                _style(ws.cell(row=row, column=self.c_utype, value=unit.unit_type))
                _style(ws.cell(row=row, column=self.c_cap, value=unit.capacity_kbd),
                       fill=FILL_INPUT, fmt="#,##0")
                _style(ws.cell(row=row, column=self.c_effcap,
                               value=self.eff_caps.get((ref.refinery_id, unit.unit_id))),
                       fill=FILL_CALC, fmt="#,##0")
                _style(ws.cell(row=row, column=self.c_out),
                       fill=FILL_OVERRIDE, fmt="0%")
                _style(ws.cell(row=row, column=self.c_mode, value="override"),
                       fill=FILL_INPUT)

                # manual override cells, prefilled from data/overrides when active
                ov_util = self.book._find_override(ref, unit, day1, "utilization")
                _style(ws.cell(row=row, column=self.c_utilov,
                               value=ov_util.value if ov_util else None),
                       fill=FILL_OVERRIDE, fmt="0.0%")
                _style(ws.cell(
                    row=row, column=self.c_util,
                    value=(f'=IF(AND({mode.format(row=row)}="override",${ovl}{row}<>""),${ovl}{row},'
                           f"SUMIFS({util_val},{util_padd},$C{row}))"
                           f"*{offx.format(row=row)}"),
                ), fill=FILL_CALC, fmt="0.0%")

                for i, cut in enumerate(self.cuts):
                    oc = get_column_letter(self._cut_ovr_col(i))
                    ov = self.book._find_override(ref, unit, day1, "yield", cut=cut)
                    _style(ws.cell(row=row, column=self._cut_ovr_col(i),
                                   value=ov.value if ov else None),
                           fill=FILL_OVERRIDE, fmt="0.0%")
                    _style(ws.cell(
                        row=row, column=self._cut_yld_col(i),
                        value=(f'=IF(AND({mode.format(row=row)}="override",${oc}{row}<>""),${oc}{row},'
                               f"SUMIFS({y_val},{y_padd},$C{row},{y_ut},$E{row},"
                               f'{y_cut},"{cut}"))'),
                    ), fill=FILL_CALC, fmt="0.0%")

                ysum = "+".join(
                    f"{get_column_letter(self._cut_yld_col(i))}{row}"
                    for i in range(len(self.cuts))
                )
                _style(ws.cell(row=row, column=self.c_ysum, value=f"={ysum}"),
                       fill=FILL_CALC, fmt="0.0%")
                cap_l, util_l = get_column_letter(self.c_cap), get_column_letter(self.c_util)
                ysum_l = get_column_letter(self.c_ysum)
                _style(ws.cell(row=row, column=self.c_base,
                               value=f"=${cap_l}{row}*${util_l}{row}*${ysum_l}{row}"),
                       fill=FILL_CALC, fmt="0.0")

                for k in range(self.weeks):
                    off_l = get_column_letter(self.c_off0 + k)
                    _style(ws.cell(row=row, column=self.c_off0 + k,
                                   value=self._offline_formula(row, k)),
                           fill=FILL_CALC, fmt="0%")
                    _style(ws.cell(
                        row=row, column=self.c_net0 + k,
                        value=(f"=${cap_l}{row}*(1-${off_l}{row})"
                               f"*${util_l}{row}*${ysum_l}{row}"),
                    ), fill=FILL_CALC, fmt="0.0")
                row += 1

            # TOTAL row
            ws.cell(row=row, column=self.c_type, value="TOTAL").font = FONT_HDR
            _style(ws.cell(row=row, column=self.c_rid, value=ref.refinery_id),
                   fill=FILL_TOTAL, bold=True)
            _style(ws.cell(row=row, column=self.c_padd, value=ref.padd),
                   fill=FILL_TOTAL, bold=True)
            _style(ws.cell(row=row, column=self.c_uid, value="NET NAPHTHA"),
                   fill=FILL_TOTAL, bold=True)
            _style(ws.cell(row=row, column=self.c_out),
                   fill=FILL_OVERRIDE, fmt="0%")
            _style(ws.cell(row=row, column=self.c_mode), fill=FILL_INPUT)
            base_l = get_column_letter(self.c_base)
            _style(ws.cell(row=row, column=self.c_base,
                           value=f"=SUM({base_l}{first_unit}:{base_l}{row - 1})"),
                   fill=FILL_TOTAL, fmt="0.0", bold=True)
            for k in range(self.weeks):
                nl = get_column_letter(self.c_net0 + k)
                _style(ws.cell(row=row, column=self.c_net0 + k,
                               value=f"=SUM({nl}{first_unit}:{nl}{row - 1})"),
                       fill=FILL_TOTAL, fmt="0.0", bold=True)
            self.refinery_total_rows[ref.refinery_id] = row
            row += 2

        widths = {1: 7, 2: 14, 3: 5, 4: 12, 5: 12, 6: 8, 7: 8, 8: 8, 9: 7}
        for i in range(len(self.cuts)):
            widths[self._cut_ovr_col(i)] = 8
            widths[self._cut_yld_col(i)] = 8
        widths[self.c_ysum] = 8
        widths[self.c_base] = 10
        for k in range(self.weeks):
            widths[self.c_net0 + k] = 7
            widths[self.c_off0 + k] = 6
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = ws.cell(row=BOX_LO, column=self.c_utilov).coordinate
        dv = DataValidation(type="list", formula1='"override,assumption"',
                            allow_blank=True)
        ws.add_data_validation(dv)
        mdl = get_column_letter(self.c_mode)
        dv.add(f"{mdl}{BOX_LO}:{mdl}{BOX_HI}")

    # -------------------------------------------------------------- Balance

    def _sheet_balance(self) -> None:
        ws = self.wb.create_sheet("Balance")
        span = 1 + self.weeks
        _banner(ws, 1, "PADD naphtha balance — kbd, weekly. + = length, - = short. "
                       "All rows are live formulas over Boxes / Flows / Demand.", span + 2)
        r = 3
        for padd in self.padds:
            _banner(ws, r, f"PADD {padd} — {PADDS.get(padd, '')}", span)
            week_row = r + 1
            _hdr(ws, week_row, 1, "Week starting")
            for k in range(self.weeks):
                cell = ws.cell(row=week_row, column=2 + k, value=f"={self.start_ref}+{k}*7")
                _style(cell, fill=FILL_CALC, fmt="m/d", bold=True)

            rows = {
                "supply": "Refinery net supply (w/ outages)",
                "base": "Refinery net supply (base, no outages)",
                "risk": "At risk — supply lost to outages",
                "flows": "Net trade flows (ships in - out)",
                "demand": "Demand (crackers, blending, diluent)",
                "balance": "BALANCE (supply + flows - demand)",
            }
            idx = {name: week_row + 1 + i for i, name in enumerate(rows)}
            for name, label in rows.items():
                bold = name == "balance"
                _style(ws.cell(row=idx[name], column=1, value=label),
                       fill=FILL_TOTAL if bold else FILL_CALC, bold=True)

            box_type = f"Boxes!$A${BOX_LO}:$A${BOX_HI}"
            box_padd = f"Boxes!$C${BOX_LO}:$C${BOX_HI}"
            base_l = get_column_letter(self.c_base)
            for k in range(self.weeks):
                col = get_column_letter(2 + k)
                wk = f"{col}${week_row}"
                net_l = get_column_letter(self.c_net0 + k)
                cur = (f'=SUMIFS(Boxes!${net_l}${BOX_LO}:${net_l}${BOX_HI},'
                       f'{box_type},"TOTAL",{box_padd},{padd})')
                base = (f'=SUMIFS(Boxes!${base_l}${BOX_LO}:${base_l}${BOX_HI},'
                        f'{box_type},"TOTAL",{box_padd},{padd})')
                flows = (f"=SUMIFS(Flows!$K${FLOW_LO}:$K${FLOW_HI},"
                         f"Flows!$B${FLOW_LO}:$B${FLOW_HI},{padd},"
                         f'Flows!$A${FLOW_LO}:$A${FLOW_HI},">="&{wk},'
                         f'Flows!$A${FLOW_LO}:$A${FLOW_HI},"<="&({wk}+6))')
                demand = (f"=SUMIFS(Demand!$C${DEM_LO}:$C${DEM_HI},"
                          f"Demand!$A${DEM_LO}:$A${DEM_HI},{padd})")
                values = {
                    "supply": cur,
                    "base": base,
                    "risk": f"={col}{idx['base']}-{col}{idx['supply']}",
                    "flows": flows,
                    "demand": demand,
                    "balance": f"={col}{idx['supply']}+{col}{idx['flows']}-{col}{idx['demand']}",
                }
                for name, formula in values.items():
                    bold = name == "balance"
                    _style(ws.cell(row=idx[name], column=2 + k, value=formula),
                           fill=FILL_TOTAL if bold else FILL_CALC, fmt="0.0", bold=bold)

            self.balance_rows[padd] = idx
            self.balance_week_row[padd] = week_row
            r = idx["balance"] + 3

        ws.column_dimensions["A"].width = 38
        for k in range(self.weeks):
            ws.column_dimensions[get_column_letter(2 + k)].width = 9
        ws.freeze_panes = "B2"

    # ---------------------------------------------------------- Calibration

    def _sheet_calibration(self) -> None:
        """Model-implied net naphtha yield vs 2024 actual, live formulas."""
        ws = self.wb.create_sheet("Calibration")
        _banner(
            ws, 1,
            "Calibration — model-implied net naphtha yield vs 2024 actuals. Big deltas "
            "mean the unit yield assumptions need tuning (or the 2024 number reflects a "
            "different naphtha definition). Yield-mode refineries match by construction.",
            10,
        )
        headers = ["refinery_id", "name", "padd", "crude kbd", "PADD util",
                   "model net kbd", "implied yield", "2024 actual", "delta (pts)", "mode"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 3, c, h)

        util_val, util_padd = self.assump_refs["util"]
        box_type = f"Boxes!$A${BOX_LO}:$A${BOX_HI}"
        box_rid = f"Boxes!$B${BOX_LO}:$B${BOX_HI}"
        base_l = get_column_letter(self.c_base)
        r = 4
        for ref in sorted(self.data.refineries, key=lambda x: (x.padd, -x.crude_capacity_kbd)):
            if ref.naphtha_yield_pct is None:
                continue
            mode = ("yield-mode" if any(u.unit_id == "CRUDE-EST" for u in ref.units)
                    else "unit-detail")
            _style(ws.cell(row=r, column=1, value=ref.refinery_id))
            _style(ws.cell(row=r, column=2, value=ref.name))
            _style(ws.cell(row=r, column=3, value=ref.padd))
            _style(ws.cell(row=r, column=4, value=ref.crude_capacity_kbd), fmt="#,##0")
            _style(ws.cell(row=r, column=5,
                           value=f"=SUMIFS({util_val},{util_padd},$C{r})"),
                   fill=FILL_CALC, fmt="0.0%")
            _style(ws.cell(
                row=r, column=6,
                value=(f'=SUMIFS(Boxes!${base_l}${BOX_LO}:${base_l}${BOX_HI},'
                       f'{box_type},"TOTAL",{box_rid},$A{r})'),
            ), fill=FILL_CALC, fmt="0.0")
            _style(ws.cell(row=r, column=7,
                           value=f'=IF($D{r}*$E{r}=0,"",$F{r}/($D{r}*$E{r}))'),
                   fill=FILL_CALC, fmt="0.00%")
            _style(ws.cell(row=r, column=8, value=ref.naphtha_yield_pct / 100.0),
                   fmt="0.00%")
            _style(ws.cell(row=r, column=9, value=f'=IF($G{r}="","",$G{r}-$H{r})'),
                   fill=FILL_CALC, fmt="+0.00%;-0.00%")
            _style(ws.cell(row=r, column=10, value=mode))
            r += 1
        last = r - 1

        # flag deltas larger than +/-2 points
        ws.conditional_formatting.add(
            f"I4:I{last}",
            CellIsRule(operator="greaterThan", formula=["0.02"], fill=FILL_FAIL))
        ws.conditional_formatting.add(
            f"I4:I{last}",
            CellIsRule(operator="lessThan", formula=["-0.02"], fill=FILL_FAIL))

        # PADD averages of 2024 actual net naphtha yield
        s = last + 3
        _banner(ws, s, "PADD simple-average 2024 net naphtha yield", 3)
        _hdr(ws, s + 1, 1, "PADD")
        _hdr(ws, s + 1, 2, "avg 2024 yield")
        _hdr(ws, s + 1, 3, "refineries")
        for i, padd in enumerate(self.padds):
            row = s + 2 + i
            _style(ws.cell(row=row, column=1, value=padd))
            _style(ws.cell(row=row, column=2,
                           value=f"=AVERAGEIF($C$4:$C${last},$A{row},$H$4:$H${last})"),
                   fill=FILL_CALC, fmt="0.00%")
            _style(ws.cell(row=row, column=3,
                           value=f"=COUNTIF($C$4:$C${last},$A{row})"),
                   fill=FILL_CALC, fmt="#,##0")

        widths = [22, 34, 6, 10, 9, 12, 12, 11, 11, 12]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A4"

    # ----------------------------------------------------------- Yields_2024

    def _sheet_yields_2024(self) -> None:
        """Full 2024 estimated-output reference table (all products)."""
        ws = self.wb.create_sheet("Yields_2024")
        _banner(ws, 1, "2024 estimated refinery yields (% of crude) — reference; "
                       "re-ingest via: python -m naphtha_model ingest-yields <file>", 13)
        from .loaders import load_yields_2024

        rows = load_yields_2024()
        headers = ["refinery_id", "padd", "state", "city", "operator", "owners",
                   "gasoil_diesel", "gasoline", "hfo", "kero_jet", "lpg",
                   "naphtha", "total"]
        keys = ["refinery_id", "padd", "state", "city", "operator", "owners",
                "gasoil_diesel_pct", "gasoline_pct", "hfo_pct", "kero_jet_pct",
                "lpg_pct", "naphtha_pct", "total_pct"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 2, c, h)
        for i, row in enumerate(rows.values(), start=3):
            for c, key in enumerate(keys, start=1):
                v = row.get(key, "")
                if key.endswith("_pct"):
                    v = float(v) / 100.0 if v not in ("", None) else None
                elif key == "padd":
                    v = int(v)
                _style(ws.cell(row=i, column=c, value=v),
                       fmt="0.00%" if key.endswith("_pct") else None)
        widths = [22, 6, 14, 16, 26, 30, 12, 10, 8, 10, 8, 10, 8]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A3"

    # ------------------------------------------------------------ Dashboard

    def _line(self, title: str, y_title: str) -> LineChart:
        ch = LineChart()
        ch.title = title
        ch.style = 2
        _axes(ch, "Week starting", y_title)
        ch.x_axis.number_format = "m/d"
        ch.height, ch.width = 8.5, 17
        return ch

    def _series(self, ws, row: int, title: str) -> Series:
        ref = Reference(ws, min_col=2, max_col=1 + self.weeks, min_row=row, max_row=row)
        s = Series(ref, title=title)
        s.marker = Marker(symbol="circle", size=5)
        s.smooth = False
        return s

    def _sheet_dashboard(self) -> None:
        ws = self.wb.create_sheet("Dashboard")
        _banner(ws, 1, "Dashboard — every chart is live: edit Assumptions / Outages / "
                       "Flows or flip the scenario toggle (Model!B5) and these move.", 20)
        bal = self.wb["Balance"]
        boxes = self.wb["Boxes"]
        padd = 3 if 3 in self.balance_rows else self.padds[0]
        idx = self.balance_rows[padd]
        cats = Reference(bal, min_col=2, max_col=1 + self.weeks,
                         min_row=self.balance_week_row[padd],
                         max_row=self.balance_week_row[padd])

        ch1 = self._line(f"PADD {padd} net naphtha supply forecast (kbd)", "kbd")
        for row_name, label in [("supply", "Supply w/ outages & scenario"),
                                ("base", "Supply base (no outages)"),
                                ("demand", "Demand")]:
            ch1.series.append(self._series(bal, idx[row_name], label))
        ch1.set_categories(cats)
        ws.add_chart(ch1, "B3")

        ch2 = self._line(f"PADD {padd} balance (+ length / - short, kbd)", "kbd")
        ch2.series.append(self._series(bal, idx["balance"], "Balance"))
        ch2.set_categories(cats)
        ws.add_chart(ch2, "L3")

        ch3 = BarChart()
        ch3.title = f"PADD {padd} supply at risk from outages (kbd)"
        ch3.y_axis.title = "kbd"
        ch3.height, ch3.width = 8.5, 17
        ch3.add_data(Reference(bal, min_col=2, max_col=1 + self.weeks,
                               min_row=idx["risk"], max_row=idx["risk"]),
                     from_rows=True, titles_from_data=False)
        ch3.series[0].tx = None
        ch3.set_categories(cats)
        _axes(ch3, "week", "kbd")
        ws.add_chart(ch3, "B21")

        ch4 = BarChart()
        ch4.title = f"PADD {padd} net cargo flows — ship tracking (kbd)"
        ch4.y_axis.title = "kbd (+in / -out)"
        ch4.height, ch4.width = 8.5, 17
        ch4.add_data(Reference(bal, min_col=2, max_col=1 + self.weeks,
                               min_row=idx["flows"], max_row=idx["flows"]),
                     from_rows=True, titles_from_data=False)
        ch4.series[0].tx = None
        ch4.set_categories(cats)
        _axes(ch4, "week", "kbd")
        ws.add_chart(ch4, "L21")

        ch5 = self._line("Net naphtha by refinery — top 10 by crude capacity (kbd/week)", "kbd")
        box_cats = Reference(boxes, min_col=self.c_net0,
                             max_col=self.c_net0 + self.weeks - 1, min_row=2, max_row=2)
        top = sorted(
            (r for r in self.data.refineries if r.crude_capacity_kbd > 0),
            key=lambda r: -r.crude_capacity_kbd,
        )[:10]
        for ref in top:
            total_row = self.refinery_total_rows[ref.refinery_id]
            ref_vals = Reference(boxes, min_col=self.c_net0,
                                 max_col=self.c_net0 + self.weeks - 1,
                                 min_row=total_row, max_row=total_row)
            s = Series(ref_vals, title=ref.name)
            s.marker = Marker(symbol="circle", size=5)
            ch5.series.append(s)
        ch5.set_categories(box_cats)
        ws.add_chart(ch5, "B39")

    # --------------------------------------------------------------- Checks

    def _sheet_checks(self) -> None:
        ws = self.wb.create_sheet("Checks")
        _banner(ws, 1, "Data checks — every check must PASS before numbers are trusted.", 6)
        _hdr(ws, 3, 1, "Check")
        _hdr(ws, 3, 2, "Failures")
        _hdr(ws, 3, 3, "Status")

        # valid-direction list used by check formulas
        for i, d in enumerate(["import", "export", "transfer_in", "transfer_out"]):
            ws.cell(row=2 + i, column=8, value=d).font = FONT_SMALL
        ws.column_dimensions["H"].hidden = True

        ref_ids = f"Refineries!$A${REF_LO + 1}:$A${REF_HI}"
        _ul = get_column_letter(self.c_util)
        yield_checks = "+".join(
            f'COUNTIF(Boxes!${get_column_letter(self._cut_yld_col(i))}${BOX_LO}:'
            f'${get_column_letter(self._cut_yld_col(i))}${BOX_HI},">1.2")'
            f'+COUNTIF(Boxes!${get_column_letter(self._cut_yld_col(i))}${BOX_LO}:'
            f'${get_column_letter(self._cut_yld_col(i))}${BOX_HI},"<-1.2")'
            for i in range(len(self.cuts))
        )
        supply_neg = "+".join(
            f'COUNTIF(Balance!$B${idx["supply"]}:'
            f'${get_column_letter(1 + self.weeks)}${idx["supply"]},"<0")'
            for idx in self.balance_rows.values()
        )
        checks = [
            ("Duplicate refinery IDs in registry",
             f"=SUMPRODUCT((COUNTIF({ref_ids},{ref_ids})>1)*({ref_ids}<>\"\"))"),
            ("Boxes units pointing at unknown refinery IDs",
             f"=SUMPRODUCT((Boxes!$A${BOX_LO}:$A${BOX_HI}=\"UNIT\")"
             f"*(COUNTIF({ref_ids},Boxes!$B${BOX_LO}:$B${BOX_HI})=0))"),
            ("Utilization outside 0-110%",
             f'=COUNTIF(Boxes!${_ul}${BOX_LO}:${_ul}${BOX_HI},">1.1")'
             f'+COUNTIF(Boxes!${_ul}${BOX_LO}:${_ul}${BOX_HI},"<0")'),
            ("Yields outside +/-120% of throughput", f"={yield_checks}"),
            ("Negative unit capacities", f'=COUNTIF(Boxes!$F${BOX_LO}:$F${BOX_HI},"<0")'),
            ("Outages ending before they start",
             f"=SUMPRODUCT((Outages!$B${OUT_LO}:$B${OUT_HI}<>\"\")"
             f"*(Outages!$F${OUT_LO}:$F${OUT_HI}<Outages!$E${OUT_LO}:$E${OUT_HI}))"),
            ("Outage offline % outside 0-100",
             f'=COUNTIF(Outages!$G${OUT_LO}:$G${OUT_HI},">100")'
             f'+COUNTIF(Outages!$G${OUT_LO}:$G${OUT_HI},"<0")'),
            ("Outages pointing at unknown refinery IDs",
             f"=SUMPRODUCT((Outages!$B${OUT_LO}:$B${OUT_HI}<>\"\")"
             f"*(COUNTIF({ref_ids},Outages!$B${OUT_LO}:$B${OUT_HI})=0))"),
            ("Flows with invalid direction",
             f"=SUMPRODUCT((Flows!$C${FLOW_LO}:$C${FLOW_HI}<>\"\")"
             f"*(COUNTIF($H$2:$H$5,Flows!$C${FLOW_LO}:$C${FLOW_HI})=0))"),
            ("Flows missing a PADD",
             f"=SUMPRODUCT((Flows!$A${FLOW_LO}:$A${FLOW_HI}<>\"\")"
             f"*(Flows!$B${FLOW_LO}:$B${FLOW_HI}=\"\"))"),
            ("Weeks where a PADD's net refinery supply is negative "
             "(reformer pull exceeds naphtha make — check yields)",
             f"={supply_neg}"),
        ]
        r = 4
        for label, formula in checks:
            ws.cell(row=r, column=1, value=label).font = FONT_SMALL
            _style(ws.cell(row=r, column=2, value=formula), fill=FILL_CALC, fmt="#,##0")
            _style(ws.cell(row=r, column=3, value=f'=IF(B{r}=0,"PASS","FAIL")'),
                   fill=FILL_CALC, bold=True)
            r += 1
        last = r - 1
        ws.cell(row=r + 1, column=1, value="OVERALL").font = Font(bold=True, size=12)
        master = ws.cell(row=r + 1, column=3,
                         value=f'=IF(SUM(B4:B{last})=0,"ALL CHECKS PASS","REVIEW FAILURES")')
        master.font = Font(bold=True, size=12)

        rng = f"C4:C{r + 1}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"FAIL"'], fill=FILL_FAIL))
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"REVIEW FAILURES"'], fill=FILL_FAIL))
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"PASS"'], fill=FILL_PASS))
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"ALL CHECKS PASS"'], fill=FILL_PASS))
        ws.column_dimensions["A"].width = 72
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 16

    # ---------------------------------------------------------------- README

    def _sheet_readme(self) -> None:
        ws = self.wb.create_sheet("README")
        _banner(ws, 1, "Zero-Based US Naphtha Model — desk workbook", 10)
        lines = [
            "",
            "WHAT THIS IS",
            "A bottom-up naphtha model: every refinery is built unit-by-unit (its 'box' on the",
            "Boxes tab), rolled up to PADD balances, with outages and ship-tracking flowing",
            "straight into the forward weekly forecast on the Dashboard.",
            "",
            "COLOR CODE",
            "", "", "",
            "HOW TO DRIVE IT",
            "1. Assumptions tab: PADD-level utilization & yields per unit type (blue).",
            "2. Boxes tab: per-unit orange override cells beat the PADD assumptions;",
            "   clear the cell to fall back. Capacity cells are blue inputs.",
            "3. Outages tab: add a row the moment an outage hits (planned or unplanned).",
            "   The weekly strip prorates partial weeks automatically.",
            "4. Scenario analysis: add rows with category = scenario, then flip",
            "   Model!B5 to YES to layer them on. Dashboard shows base vs case.",
            "5. Flows tab: ship-tracking cargoes (imports/exports/transfers) hit the",
            "   balance in the week of their date.",
            "6. Checks tab: must read ALL CHECKS PASS before trusting the numbers.",
            "",
            "MODEL CUT",
            "Net naphtha = sum over units of capacity x (1 - offline%) x utilization x yield.",
            "Yields are signed: reformers/isom consume naphtha (negative), so the total is",
            "net naphtha available to the market.",
            "",
            "PLACEHOLDER WARNING",
            "Registry, capacities, yields and demand are illustrative placeholders until the",
            "real capacity sheet and desk assumptions are loaded. Do not trade off them yet.",
            "",
            "Regenerate from the data/ directory:  python -m naphtha_model export --out <file>",
        ]
        for i, text in enumerate(lines, start=2):
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = FONT_HDR if text.isupper() and text else FONT_SMALL
        # color legend swatches
        legend = [(FILL_INPUT, "BLUE — input: type here"),
                  (FILL_OVERRIDE, "ORANGE — manual override: beats assumptions; clear to revert"),
                  (FILL_CALC, "GREY — formula: do not type")]
        for j, (fill, label) in enumerate(legend):
            row = 9 + j
            ws.cell(row=row, column=2).fill = fill
            ws.cell(row=row, column=2).border = BORDER
            ws.cell(row=row, column=3, value=label).font = FONT_SMALL
        ws.column_dimensions["A"].width = 90


class SimpleWorkbook(DeskWorkbook):
    """The desk workbook, one tab per methodology step:

    Data        imported registry + 2024 yields (+ crude capacity inputs)
    Boxes       every refinery, unit by unit (nameplate, eff cap, actual
                utilization, yields, net naphtha)
    Assumptions per-PADD utilization/yield inputs, yield-mode cut split,
                plus the logistics & blending-economics input blocks
    Nameplate   stated unit capacities pivoted refinery x unit (live from
                Boxes)
    Effective   demonstrated capacities (max annual throughput 2017-2024
                excl. 2020) with PADD nameplate/effective/running rollup
    CrudeSlate  actual crude diet per refinery + purchased feedstocks
                (incl. merchant naphtha buyers)
    BlendEcon   blend value spreads, arbs, max-light/max-heavy slate
                scenarios - driven by the Assumptions price/freight inputs
    KitWalk     the naphtha path per refinery: CDU -> SR naphtha -> NHT ->
                reformer -> net, against the 2024 actual (yield tuning)
    """

    TAB_ORDER = ["Cover", "Assumptions", "Individual Refineries", "Outages",
                 "Outage History", "Kpler Flows", "US Balance", "Live Feeds",
                 "CrudeSlate", "BlendEcon", "KitWalk", "Nameplate",
                 "Effective", "Data"]

    def build(self, out_path: Path) -> Path:
        # Boxes sheet extent is deterministic; grids on Assumptions need it
        # before the boxes are written (banner + units + total + blank per box)
        self.box_last = 2 + sum(len(r.units) + 3 for r in self.data.refineries)
        self._sheet_assumptions()
        self._simple_share_block()
        self._econ_assumption_blocks()
        self._sheet_data()
        self._simple_boxes()
        self._sheet_nameplate()
        self._sheet_effective()
        self._sheet_outages_tab()
        self._sheet_outage_history()
        self._sheet_kpler()
        self._sheet_usbalance()
        self._sheet_livefeeds()
        self._sheet_crudeslate()
        self._sheet_blendecon()
        self._sheet_kitwalk()
        self._sheet_cover()
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        self.wb._sheets = [self.wb[n] for n in self.TAB_ORDER]
        for name in self.TAB_ORDER:
            ws = self.wb[name]
            _theme(ws, TAB_COLORS.get(name))
            # clean printing: landscape, fit to one page wide
            from openpyxl.worksheet.properties import PageSetupProperties

            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        # sort/filter on the grid tabs
        filters = {
            "Data": f"A2:O{self.data_last_row}",
            "Nameplate": f"A2:R{self.nameplate_last}",
            "Effective": f"A2:T{2 + len(self.data.refineries)}",
            "KitWalk": f"A2:O{2 + sum(1 for r in self.data.refineries if r.units and r.units[0].unit_id != 'CRUDE-EST')}",
        }
        for name, ref in filters.items():
            self.wb[name].auto_filter.ref = ref
        out_path = Path(out_path)
        self.wb.save(out_path)
        return out_path

    # ------------------------------------------------------------ Cover tab

    def _sheet_cover(self) -> None:
        """The main dashboard: tradable takeaways — supply, outages, balance,
        spreads/arbs — all live formulas + four charts, no digging required."""
        from datetime import datetime as _dt

        from openpyxl.chart import BarChart, LineChart, Reference, Series

        ws = self.wb.create_sheet("Cover")
        for c in range(1, 16):
            ws.column_dimensions[get_column_letter(c)].width = 11
        ws.column_dimensions["A"].width = 3

        # masthead
        for r in range(2, 7):
            for c in range(2, 16):
                ws.cell(row=r, column=c).fill = FILL_BANNER
        for c in range(2, 16):
            ws.cell(row=7, column=c).fill = FILL_GOLD
        ws.row_dimensions[7].height = 4
        t = ws.cell(row=3, column=3, value="US NAPHTHA — ZERO-BASED REFINERY MODEL")
        t.font = Font(bold=True, size=20, color="FFFFFF", name="Calibri")
        st = ws.cell(
            row=5, column=3,
            value=f"{len(self.data.refineries)} refineries · unit-level build · "
                  f"PADD 1–5 · generated {_dt.now():%d %b %Y}",
        )
        st.font = Font(size=10, color="C9A227", name="Calibri")

        # live KPI band
        base_l = get_column_letter(self.c_base)
        utl_l = get_column_letter(self.c_util)
        BOX = "'Individual Refineries'"
        b_typ = f"{BOX}!$A$3:$A${self.box_last}"
        b_pad = f"{BOX}!$C$3:$C${self.box_last}"
        b_uid = f"{BOX}!$D$3:$D${self.box_last}"
        b_cap = f"{BOX}!$F$3:$F${self.box_last}"
        b_utl = f"{BOX}!${utl_l}$3:${utl_l}${self.box_last}"
        b_net = f"{BOX}!${base_l}$3:${base_l}${self.box_last}"
        kpis = [
            ("US net naphtha (kbd)",
             f'=SUMPRODUCT(({b_typ}="TOTAL")*{b_net})', "#,##0"),
            ("PADD 3 share",
             f'=SUMPRODUCT(({b_typ}="TOTAL")*({b_pad}=3)*{b_net})'
             f'/SUMPRODUCT(({b_typ}="TOTAL")*{b_net})', "0%"),
            ("CDU offline this wk (kbd)", "=Outages!$B$5", "#,##0"),
            ("US crude runs (kbd)",
             f'=SUMPRODUCT((({b_uid}="CDU")+({b_uid}="CRUDE-EST"))*{b_cap}*{b_utl})',
             "#,##0"),
            ("EA balance, latest mo (kbd)",
             f"='US Balance'!$I${self.usbal_last}", "+#,##0;-#,##0"),
            ("units modelled",
             f'=SUMPRODUCT(({b_typ}="UNIT")*1)', "#,##0"),
        ]
        for i, (label, formula, fmt) in enumerate(kpis):
            col = 3 + i * 2
            v = ws.cell(row=9, column=col, value=formula)
            v.font = Font(bold=True, size=16, color=NAVY, name="Calibri")
            v.number_format = fmt
            v.border = Border(top=GOLD_EDGE)
            cap = ws.cell(row=10, column=col, value=label.upper())
            cap.font = Font(size=7, color="7A8699", name="Calibri")

        # chart feeds: net naphtha by PADD
        for i, p in enumerate(self.padds):
            lbl = ws.cell(row=60 + i, column=18, value=f"PADD {p}")
            val = ws.cell(row=60 + i, column=19,
                          value=f'=SUMPRODUCT(({b_typ}="TOTAL")*({b_pad}={p})*{b_net})')
            lbl.font = val.font = Font(size=8, color="AAAAAA")

        ch1 = BarChart()
        ch1.type = "col"
        ch1.title = "Net naphtha by PADD (kbd)"
        ch1.height, ch1.width = 7.5, 11
        ch1.legend = None
        ser = Series(Reference(ws, min_col=19, min_row=60, max_row=59 + len(self.padds)))
        ser.graphicalProperties.solidFill = GOLD
        ch1.series.append(ser)
        ch1.set_categories(Reference(ws, min_col=18, min_row=60,
                                     max_row=59 + len(self.padds)))
        _axes(ch1, "PADD", "kbd")
        ws.add_chart(ch1, "C12")

        ot = self.wb["Outages"]
        ch2 = BarChart()
        ch2.type = "col"
        ch2.title = "CDU capacity offline — next 13 weeks (kbd)"
        ch2.height, ch2.width = 7.5, 12
        ch2.legend = None
        ser = Series(Reference(ot, min_col=2, max_col=14, min_row=5, max_row=5))
        ser.graphicalProperties.solidFill = "B03A2E"
        ch2.series.append(ser)
        ch2.set_categories(Reference(ot, min_col=2, max_col=14, min_row=4, max_row=4))
        _axes(ch2, "week starting", "kbd offline", rotate=True)
        ws.add_chart(ch2, "I12")

        ub = self.wb["US Balance"]
        ch3 = LineChart()
        ch3.title = "US naphtha imports / exports / balance (kbd, monthly)"
        ch3.height, ch3.width = 7.5, 11
        for colx, lab, color in ((3, "imports", "1F6FB2"), (4, "exports", "B03A2E"),
                                 (9, "balance", GOLD)):
            ser = Series(Reference(ub, min_col=colx, min_row=self.usbal_first,
                                   max_row=self.usbal_last), title=lab)
            ser.graphicalProperties.line.solidFill = color
            ch3.series.append(ser)
        ch3.set_categories(Reference(ub, min_col=1, min_row=self.usbal_first,
                                     max_row=self.usbal_last))
        _axes(ch3, "month", "kbd", rotate=True, skip=3)
        ws.add_chart(ch3, "C28")

        eff = self.wb["Effective"]
        ch4 = BarChart()
        ch4.type = "col"
        ch4.title = "CDU by PADD: nameplate vs effective vs running (kbd)"
        ch4.height, ch4.width = 7.5, 12
        r0 = self.eff_rollup_first
        r1 = r0 + len(self.padds) - 1
        for colx, label, color in ((2, "nameplate", NAVY), (3, "effective", "5B7BA8"),
                                   (4, "running", GOLD)):
            ser = Series(Reference(eff, min_col=colx, min_row=r0, max_row=r1),
                         title=label)
            ser.graphicalProperties.solidFill = color
            ch4.series.append(ser)
        ch4.set_categories(Reference(eff, min_col=1, min_row=r0, max_row=r1))
        _axes(ch4, "PADD", "kbd")
        ws.add_chart(ch4, "I28")

        # THE TAPE — tradable takeaways, all live
        _banner(ws, 44, "  THE TAPE — live takeaways", 14)
        _hdr(ws, 45, 3, "PADD")
        _hdr(ws, 45, 4, "net naphtha kbd")
        _hdr(ws, 45, 5, "share")
        for i, p in enumerate(self.padds):
            r = 46 + i
            _style(ws.cell(row=r, column=3, value=p))
            _style(ws.cell(
                row=r, column=4,
                value=f'=SUMPRODUCT(({b_typ}="TOTAL")*({b_pad}={p})*{b_net})',
            ), fill=FILL_CALC, fmt="#,##0.0")
            _style(ws.cell(
                row=r, column=5,
                value=f'=$D{r}/SUMPRODUCT(({b_typ}="TOTAL")*{b_net})',
            ), fill=FILL_CALC, fmt="0%")

        _hdr(ws, 45, 7, "spread / arb")
        _hdr(ws, 45, 8, "$/bbl")
        _hdr(ws, 45, 9, "signal")
        tape = [
            ("Gasoline - naphtha", "='BlendEcon'!$D$4",
             '=IF(\'BlendEcon\'!$D$4>0,"naphtha into mogas","")'),
            ("Reformate uplift", "='BlendEcon'!$D$6", ""),
            ("Naphtha crack (vs WTI)", "='BlendEcon'!$D$7", ""),
            ("USGC -> Asia arb", "='BlendEcon'!$D$10", "='BlendEcon'!$E$10"),
            ("USGC -> NWE arb", "='BlendEcon'!$D$11", "='BlendEcon'!$E$11"),
            ("NWE -> USGC arb", "='BlendEcon'!$D$12", "='BlendEcon'!$E$12"),
        ]
        for i, (label, val, sig) in enumerate(tape):
            r = 46 + i
            _style(ws.cell(row=r, column=7, value=label))
            _style(ws.cell(row=r, column=8, value=val), fill=FILL_CALC,
                   fmt="+0.00;-0.00")
            _style(ws.cell(row=r, column=9, value=sig or None), fill=FILL_CALC)

        _hdr(ws, 45, 11, "balance nowcast")
        _hdr(ws, 45, 12, "kbd")
        now = [
            ("Model supply", "='US Balance'!$C$4"),
            ("+ Kpler imports", "='US Balance'!$C$5"),
            ("- Kpler exports", "='US Balance'!$C$6"),
            ("- Demand", "='US Balance'!$C$7"),
            ("IMPLIED BALANCE", "='US Balance'!$C$8"),
        ]
        for i, (label, val) in enumerate(now):
            r = 46 + i
            _style(ws.cell(row=r, column=11, value=label), bold=(i == 4))
            _style(ws.cell(row=r, column=12, value=val),
                   fill=FILL_TOTAL if i == 4 else FILL_CALC,
                   fmt="+#,##0.0;-#,##0.0", bold=(i == 4))

        # model map
        _banner(ws, 54, "  MODEL MAP", 14)
        guide = [
            ("Data", "imported registry, 2024 net yields, crude capacities (inputs)"),
            ("Individual Refineries", "the engine — every refinery unit by unit; net naphtha per site"),
            ("Assumptions", "PADD yields & utilization, dials, BBG prices, manual freight"),
            ("Outages", "current & planned events (Snowflake landing) + weekly at-risk strip"),
            ("Outage History", "2023+ history, TAR seasonality, planned vs unplanned stats"),
            ("Kpler Flows", "ship tracking: trades / fixtures / flows by grade & status"),
            ("US Balance", "EA monthly actuals + live balance nowcast"),
            ("Live Feeds", "in-workbook API pulls: EIA weekly/monthly, IIR events"),
            ("CrudeSlate", "actual crude diet; merchant naphtha buyers flagged"),
            ("BlendEcon", "spreads, arbs, max-light vs max-heavy — with charts"),
            ("KitWalk", "CDU -> SR naphtha -> NHT -> reformer -> net, vs 2024 actuals"),
            ("Nameplate", "stated unit capacities, refinery x unit"),
            ("Effective", "demonstrated capacities vs nameplate vs running"),
        ]
        for i, (name, desc) in enumerate(guide):
            r = 56 + i
            n = ws.cell(row=r, column=3, value=name)
            n.font = Font(bold=True, size=9, color="1F5AA8",
                          underline="single", name="Calibri")
            n.hyperlink = f"#'{name}'!A1"
            d = ws.cell(row=r, column=6, value=desc)
            d.font = FONT_SMALL

        # legend + finder
        _banner(ws, 70, "  HOW TO DRIVE IT", 14)
        legend = [(FILL_INPUT, "blue — input: type here"),
                  (FILL_OVERRIDE, "orange — manual override/outage: applies when MODE = override"),
                  (FILL_CALC, "grey — live formula: don't type")]
        for i, (fill, text) in enumerate(legend):
            r = 72 + i
            sw = ws.cell(row=r, column=3)
            sw.fill = fill
            sw.border = BORDER
            ws.cell(row=r, column=4, value=text).font = FONT_SMALL

        _banner(ws, 77, "  REFINERY FINDER", 14)
        pick = ws.cell(row=79, column=3, value="MOTIVA_PAR")
        _style(pick, fill=FILL_INPUT)
        dv = DataValidation(type="list",
                            formula1=f"=Data!$A$3:$A${self.data_last_row}",
                            allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("C79")
        d_id = f"Data!$A$3:$A${self.data_last_row}"
        finder = [
            ("name", f'=IFERROR(INDEX(Data!$B$3:$B${self.data_last_row},'
                     f'MATCH($C$79,{d_id},0)),"")', None),
            ("PADD", f'=IFERROR(INDEX(Data!$D$3:$D${self.data_last_row},'
                     f'MATCH($C$79,{d_id},0)),"")', None),
            ("crude kbd", f"=SUMIFS(Data!$G$3:$G${self.data_last_row},{d_id},$C$79)",
             "#,##0"),
            ("net naphtha kbd",
             f'=SUMPRODUCT(({b_typ}="TOTAL")*({BOX}!$B$3:$B${self.box_last}=$C$79)'
             f"*{b_net})", "0.0"),
            ("2024 actual naphtha %",
             f"=SUMIFS(Data!$N$3:$N${self.data_last_row},{d_id},$C$79)", "0.00%"),
        ]
        for i, (label, formula, fmt) in enumerate(finder):
            rr = 81 + i
            ws.cell(row=rr, column=3, value=label).font = FONT_SMALL
            cell = ws.cell(row=rr, column=5, value=formula)
            cell.font = Font(bold=True, size=10, color=NAVY, name="Calibri")
            if fmt:
                cell.number_format = fmt
        jump = ws.cell(row=79, column=5, value=(
            "=HYPERLINK(\"#'Individual Refineries'!A\""
            "&MATCH($C$79,'Individual Refineries'!$B$1:$B$3000,0)-1,"
            '"→ open this refinery\'s box")'
        ))
        jump.font = Font(size=10, color="1F5AA8", underline="single", name="Calibri")

    def _simple_share_block(self) -> None:
        """Yield-mode cut split inputs, placed right of the lookup tables."""
        ws = self.wb["Assumptions"]
        ws["A1"] = (
            "Assumptions — blue cells are inputs. Yields are % of unit throughput; "
            "consumers (reformer/isom) are NEGATIVE. Edits flow straight into the Individual Refineries sheet."
        )
        col = 6 + len(self.cuts) + 10
        _hdr(ws, 3, col, "Yield-mode cut split")
        _hdr(ws, 3, col + 1, "share")
        shares = (self.book.global_cfg.get("yield_mode") or {}).get("cut_shares") or {}
        self.share_refs: dict[str, str] = {}
        for i, cut in enumerate(self.cuts):
            row = 4 + i
            ws.cell(row=row, column=col, value=cut).font = FONT_SMALL
            cell = ws.cell(row=row, column=col + 1, value=float(shares.get(cut, 0.0)))
            _style(cell, fill=FILL_INPUT, fmt="0%")
            self.share_refs[cut] = f"Assumptions!${get_column_letter(col + 1)}${row}"
        note = ws.cell(row=4 + len(self.cuts) + 1, column=col,
                       value="how a yield-mode refinery's 2024 net yield splits across cuts")
        note.font = FONT_SMALL

    def _sheet_data(self) -> None:
        from .loaders import load_yields_2024

        ws = self.wb.create_sheet("Data")
        _banner(
            ws, 1,
            "Imported data — refinery registry + 2024 net yields (% of crude). "
            "Type a crude capacity (blue) and that refinery's box lights up. "
            "Re-import: python -m naphtha_model ingest-yields <file>",
            15,
        )
        headers = ["refinery_id", "name", "owner", "padd", "state", "city",
                   "crude_capacity_kbd", "status", "gasoil_diesel", "gasoline",
                   "hfo", "kero_jet", "lpg", "naphtha", "total"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 2, c, h)
        yields = load_yields_2024()
        y_keys = ["gasoil_diesel_pct", "gasoline_pct", "hfo_pct", "kero_jet_pct",
                  "lpg_pct", "naphtha_pct", "total_pct"]
        r = 3
        for ref in sorted(self.data.refineries,
                          key=lambda x: (x.padd, -x.crude_capacity_kbd, x.name)):
            y = yields.get(ref.refinery_id, {})
            _style(ws.cell(row=r, column=1, value=ref.refinery_id))
            _style(ws.cell(row=r, column=2, value=ref.name))
            _style(ws.cell(row=r, column=3, value=ref.owner))
            _style(ws.cell(row=r, column=4, value=ref.padd))
            _style(ws.cell(row=r, column=5, value=ref.state))
            _style(ws.cell(row=r, column=6, value=ref.city))
            _style(ws.cell(row=r, column=7, value=ref.crude_capacity_kbd),
                   fill=FILL_INPUT, fmt="#,##0")
            _style(ws.cell(row=r, column=8, value=ref.status))
            for j, key in enumerate(y_keys):
                v = y.get(key)
                _style(ws.cell(row=r, column=9 + j,
                               value=float(v) / 100.0 if v not in (None, "") else None),
                       fill=FILL_INPUT, fmt="0.00%")
            r += 1
        self.data_last_row = r - 1
        widths = [22, 30, 30, 6, 13, 15, 10, 9, 12, 9, 8, 9, 8, 9, 8]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A3"

    def _simple_boxes(self) -> None:
        ws = self.wb.create_sheet("Individual Refineries")
        last_col = self.c_base
        _banner(
            ws, 1,
            "Refinery boxes — net naphtha = capacity x utilization x signed yields. "
            "Blue = input, orange = manual override, grey = formula. MODE flips a "
            "unit between override and assumption; OFFLINE % knocks capacity out "
            "(e.g. 1 = full outage, 0.5 = half rates). Set either on the NET "
            "NAPHTHA row to apply refinery-wide. Yield-mode rows read Data.",
            last_col,
        )
        labels = {
            self.c_type: "row", self.c_rid: "refinery_id", self.c_padd: "padd",
            self.c_uid: "unit", self.c_utype: "type", self.c_cap: "cap kbd",
            self.c_effcap: "eff cap", self.c_out: "offline %",
            self.c_mode: "mode", self.c_utilov: "util ovr", self.c_util: "util",
            self.c_ysum: "Σ yield", self.c_base: "net naphtha kbd",
        }
        for i, cut in enumerate(self.cuts):
            labels[self._cut_ovr_col(i)] = f"{cut} ovr"
            labels[self._cut_yld_col(i)] = f"{cut} yield"
        for col, text in labels.items():
            _hdr(ws, 2, col, text)

        util_val, util_padd = self.assump_refs["util"]
        y_val, y_padd, y_ut, y_cut = self.assump_refs["yield"]
        data_id = f"Data!$A$3:$A${self.data_last_row}"
        data_cap = f"Data!$G$3:$G${self.data_last_row}"
        data_naph = f"Data!$N$3:$N${self.data_last_row}"
        util_dial = self.dial_refs["Utilization scaler (100% = as-is)"]
        yld_dial = self.dial_refs["Naphtha yield scaler (100% = as-is)"]
        # Effective-tab grid geometry (built after Boxes; layout is fixed)
        eff_last = 2 + len(self.data.refineries)
        eff_c0 = get_column_letter(5)
        eff_c1 = get_column_letter(4 + len(self.UNIT_COLS))
        day1 = self.axis[0]
        # live outage snapshot (desk offline-events export), if ingested
        cur_out: dict[tuple[str, str], float] = {}
        out_path = DATA_DIR / "reference" / "current_outages.csv"
        if out_path.exists():
            import csv as _csv

            with out_path.open() as fh:
                for orow in _csv.DictReader(fh):
                    cur_out[(orow["refinery_id"], orow["unit_id"])] = float(
                        orow["offline_frac"])
        row = 3
        self.box_banner_rows: dict[str, int] = {}
        for ref in sorted(self.data.refineries,
                          key=lambda x: (x.padd, -x.crude_capacity_kbd, x.name)):
            self.box_banner_rows[ref.refinery_id] = row
            _banner(ws, row, "", last_col)
            name = ref.name.replace('"', "'")
            owner = ref.owner.replace('"', "'")
            ws.cell(row=row, column=1).value = (
                f'="  {name}  [{ref.refinery_id}]   —   {owner}   —   '
                f'PADD {ref.padd}   —   crude "&TEXT(SUMIFS({data_cap},{data_id},'
                f'"{ref.refinery_id}"),"#,##0")&" kbd"'
            )
            row += 1
            first_unit = row
            ovl = get_column_letter(self.c_utilov)
            mdl = get_column_letter(self.c_mode)
            outl = get_column_letter(self.c_out)
            total_row = first_unit + len(ref.units)
            offx = (f"(1-MIN(1,MAX(N(${outl}{{row}}),"
                    f"N(${outl}{total_row}))))")
            mode = f'IF(${mdl}{total_row}<>"",${mdl}{total_row},${mdl}{{row}})'
            for unit in ref.units:
                est = unit.unit_id == "CRUDE-EST"
                ws.cell(row=row, column=self.c_type, value="UNIT").font = FONT_SMALL
                _style(ws.cell(row=row, column=self.c_rid, value=ref.refinery_id))
                _style(ws.cell(row=row, column=self.c_padd, value=ref.padd))
                _style(ws.cell(row=row, column=self.c_uid, value=unit.unit_id))
                _style(ws.cell(row=row, column=self.c_utype,
                               value="est. yield" if est else unit.unit_type))
                if est:
                    _style(ws.cell(row=row, column=self.c_cap,
                                   value=f"=SUMIFS({data_cap},{data_id},$B{row})"),
                           fill=FILL_CALC, fmt="#,##0")
                else:
                    _style(ws.cell(row=row, column=self.c_cap, value=unit.capacity_kbd),
                           fill=FILL_INPUT, fmt="#,##0")
                eff_inner = (
                    f"INDEX(Effective!${eff_c0}$3:${eff_c1}${eff_last},"
                    f"MATCH($B{row},Effective!$A$3:$A${eff_last},0),"
                    f"MATCH($D{row},Effective!${eff_c0}$2:${eff_c1}$2,0))"
                )
                _style(ws.cell(
                    row=row, column=self.c_effcap,
                    value=f'=IFERROR(IF({eff_inner}=0,"",{eff_inner}),"")',
                ), fill=FILL_CALC, fmt="#,##0")
                _style(ws.cell(row=row, column=self.c_out,
                               value=cur_out.get((ref.refinery_id, unit.unit_id))),
                       fill=FILL_OVERRIDE, fmt="0%")
                _style(ws.cell(row=row, column=self.c_mode, value="assumption"),
                       fill=FILL_INPUT)

                ov_util = self.book._find_override(ref, unit, day1, "utilization")
                _style(ws.cell(row=row, column=self.c_utilov,
                               value=ov_util.value if ov_util else None),
                       fill=FILL_OVERRIDE, fmt="0.0%")
                _style(ws.cell(
                    row=row, column=self.c_util,
                    value=(f'=IF(AND({mode.format(row=row)}="override",${ovl}{row}<>""),${ovl}{row},'
                           f"SUMIFS({util_val},{util_padd},$C{row}))*{util_dial}"
                           f"*{offx.format(row=row)}"),
                ), fill=FILL_CALC, fmt="0.0%")

                for i, cut in enumerate(self.cuts):
                    oc = get_column_letter(self._cut_ovr_col(i))
                    if est:
                        # 2024 net yield (Data sheet) split by the cut shares
                        _style(ws.cell(row=row, column=self._cut_ovr_col(i)),
                               fill=FILL_OVERRIDE, fmt="0.0%")
                        formula = (
                            f'=IF(AND({mode.format(row=row)}="override",${oc}{row}<>""),${oc}{row},'
                            f"SUMIFS({data_naph},{data_id},$B{row})"
                            f"*{self.share_refs[cut]})"
                        )
                    else:
                        ov = self.book._find_override(ref, unit, day1, "yield", cut=cut)
                        _style(ws.cell(row=row, column=self._cut_ovr_col(i),
                                       value=ov.value if ov else None),
                               fill=FILL_OVERRIDE, fmt="0.0%")
                        formula = (
                            f'=IF(AND({mode.format(row=row)}="override",${oc}{row}<>""),${oc}{row},'
                            f"SUMIFS({y_val},{y_padd},$C{row},{y_ut},$E{row},"
                            f'{y_cut},"{cut}"))'
                        )
                    _style(ws.cell(row=row, column=self._cut_yld_col(i), value=formula),
                           fill=FILL_CALC, fmt="0.00%")

                ysum = "+".join(
                    f"{get_column_letter(self._cut_yld_col(i))}{row}"
                    for i in range(len(self.cuts))
                )
                _style(ws.cell(row=row, column=self.c_ysum,
                               value=f"=({ysum})*{yld_dial}"),
                       fill=FILL_CALC, fmt="0.00%")
                cap_l = get_column_letter(self.c_cap)
                util_l = get_column_letter(self.c_util)
                ysum_l = get_column_letter(self.c_ysum)
                _style(ws.cell(row=row, column=self.c_base,
                               value=f"=${cap_l}{row}*${util_l}{row}*${ysum_l}{row}"),
                       fill=FILL_CALC, fmt="0.0")
                row += 1

            ws.cell(row=row, column=self.c_type, value="TOTAL").font = FONT_HDR
            _style(ws.cell(row=row, column=self.c_rid, value=ref.refinery_id),
                   fill=FILL_TOTAL, bold=True)
            _style(ws.cell(row=row, column=self.c_padd, value=ref.padd),
                   fill=FILL_TOTAL, bold=True)
            _style(ws.cell(row=row, column=self.c_uid, value="NET NAPHTHA"),
                   fill=FILL_TOTAL, bold=True)
            _style(ws.cell(row=row, column=self.c_out),
                   fill=FILL_OVERRIDE, fmt="0%")
            _style(ws.cell(row=row, column=self.c_mode), fill=FILL_INPUT)
            base_l = get_column_letter(self.c_base)
            _style(ws.cell(row=row, column=self.c_base,
                           value=f"=SUM({base_l}{first_unit}:{base_l}{row - 1})"),
                   fill=FILL_TOTAL, fmt="0.0", bold=True)
            self.refinery_total_rows[ref.refinery_id] = row
            row += 2

        widths = {1: 7, 2: 22, 3: 5, 4: 12, 5: 12, 6: 8, 7: 8,
                  self.c_out: 8, self.c_mode: 11, self.c_utilov: 8,
                  self.c_util: 7}
        for i in range(len(self.cuts)):
            widths[self._cut_ovr_col(i)] = 8
            widths[self._cut_yld_col(i)] = 9
        widths[self.c_ysum] = 9
        widths[self.c_base] = 14
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A3"
        self.box_last = row - 1
        # per-unit / per-refinery override-vs-assumption dropdown
        dv = DataValidation(type="list", formula1='"override,assumption"',
                            allow_blank=True)
        ws.add_data_validation(dv)
        mdl = get_column_letter(self.c_mode)
        dv.add(f"{mdl}3:{mdl}{self.box_last}")

    # ----------------------------------------------- Assumptions: econ blocks

    def _econ_assumption_blocks(self) -> None:
        """Master dials plus the desk's market-data block (prices with
        Bloomberg BDP wiring, freight, blend scenarios) from
        data/assumptions/market.yaml — layout: label | value | ticker | note."""
        import yaml as _yaml

        ws = self.wb["Assumptions"]
        col = 6 + len(self.cuts) + 10           # same column as the share block
        cl, vl = get_column_letter(col), get_column_letter(col + 1)
        tl, nl = get_column_letter(col + 2), get_column_letter(col + 3)
        with (DATA_DIR / "assumptions" / "market.yaml").open() as fh:
            mkt = _yaml.safe_load(fh)
        r = 4 + len(self.cuts) + 3

        # master dials
        _hdr(ws, r, col, "MASTER DIALS — one cell flexes the whole model")
        _hdr(ws, r, col + 1, "x")
        self.dial_refs = {}
        for label, val in [("Utilization scaler (100% = as-is)", 1.0),
                           ("Naphtha yield scaler (100% = as-is)", 1.0)]:
            r += 1
            ws.cell(row=r, column=col, value=label).font = FONT_SMALL
            _style(ws.cell(row=r, column=col + 1, value=val),
                   fill=FILL_INPUT, fmt="0%")
            self.dial_refs[label] = f"Assumptions!${vl}${r}"
        r += 2

        def market_block(title, hdr_val, rows, fmt, codes_hdr="notes"):
            nonlocal r
            _hdr(ws, r, col, title)
            _hdr(ws, r, col + 1, hdr_val)
            _hdr(ws, r, col + 2, codes_hdr)
            hdr_row = r
            refs = {}
            for spec in rows:
                r += 1
                ws.cell(row=r, column=col, value=spec["label"]).font = FONT_SMALL
                source = spec.get("source", "static")
                if source == "bdp":
                    value = f"=_xll.BDP(${tl}${r}, ${vl}${hdr_row})"
                elif source == "formula":
                    value = spec["formula"].replace("{next}", str(r + 1)) \
                                           .replace("S", vl)
                else:
                    value = spec.get("value")
                _style(ws.cell(row=r, column=col + 1, value=value),
                       fill=FILL_INPUT, fmt=fmt)
                if spec.get("ticker"):
                    _style(ws.cell(row=r, column=col + 2, value=spec["ticker"]),
                           fill=FILL_INPUT)
                if spec.get("note"):
                    # notes sit in the ticker column when there is no ticker
                    # (matches the desk's layout), else one column further
                    note_col = col + 3 if spec.get("ticker") else col + 2
                    ws.cell(row=r, column=note_col,
                            value=spec["note"]).font = FONT_SMALL
                refs[spec["key"]] = f"Assumptions!${vl}${r}"
            r += 2
            return refs

        self.price_refs = market_block(
            "Prices — live via Bloomberg BDP (needs terminal)",
            mkt.get("header", "PX_LAST"), mkt["prices"], "0.00",
            codes_hdr="BBG Codes")

        # ---- freight: entered AS QUOTED, converted to $/bbl for the model
        _hdr(ws, r, col, "Logistics / freight — MANUAL entries (quoted rates)")
        _hdr(ws, r, col + 1, "rate as quoted")
        _hdr(ws, r, col + 2, "unit")
        _hdr(ws, r, col + 3, "$/bbl used")
        hdr_row = r
        r += 1
        ws.cell(row=r, column=col,
                value="Naphtha conversion (bbl per mt)").font = FONT_SMALL
        _style(ws.cell(row=r, column=col + 1,
                       value=float(mkt.get("bbl_per_mt", 8.9))),
               fill=FILL_INPUT, fmt="0.0")
        bblmt = f"${vl}${r}"
        self.freight_refs = {}
        dvu = DataValidation(type="list", formula1='"$/mt,$/bbl"',
                             allow_blank=False)
        ws.add_data_validation(dvu)
        for spec in mkt["freight"]:
            r += 1
            ws.cell(row=r, column=col, value=spec["label"]).font = FONT_SMALL
            if spec.get("source") == "formula":
                quoted = spec["formula"].replace("{next}", str(r + 1)) \
                                        .replace("S", vl)
            else:
                quoted = spec.get("value")
            _style(ws.cell(row=r, column=col + 1, value=quoted),
                   fill=FILL_INPUT, fmt="0.00")
            _style(ws.cell(row=r, column=col + 2,
                           value=spec.get("quote_unit", "$/bbl")),
                   fill=FILL_INPUT)
            dvu.add(f"{tl}{r}")
            _style(ws.cell(
                row=r, column=col + 3,
                value=(f'=IF(${tl}{r}="$/mt",${vl}{r}/{bblmt},${vl}{r})'),
            ), fill=FILL_CALC, fmt="0.00")
            if spec.get("note"):
                ws.cell(row=r, column=col + 4,
                        value=spec["note"]).font = FONT_SMALL
            self.freight_refs[spec["key"]] = f"Assumptions!${nl}${r}"
        r += 2

        self.blend_refs = market_block(
            "Blend scenario assumptions — MANUAL entries", "value",
            mkt["blend_scenarios"], "0.00")

        # ---- demand: one number feeds the balance; three ways to set it
        import csv as _csv

        bal_path = DATA_DIR / "reference" / "us_naphtha_balance_monthly.csv"
        months: list[str] = []
        last_dem = None
        if bal_path.exists():
            rows = list(_csv.DictReader(bal_path.open()))
            months = sorted({row["month"] for row in rows})
            dem = {row["month"]: float(row["kbd"] or 0)
                   for row in rows if row["flow"] == "TOTDEMO"}
            if months:
                last_dem = dem.get(months[-1])
        # EA demand column on US Balance: col E, data rows start at 12
        ea_rng = f"'US Balance'!$E$12:$E${11 + len(months)}"
        _hdr(ws, r, col, "DEMAND — US naphtha demand used by the balance (kbd)")
        _hdr(ws, r, col + 1, "value")
        r += 1
        ws.cell(row=r, column=col,
                value="Source (Manual / EA latest / EIA feed)").font = FONT_SMALL
        _style(ws.cell(row=r, column=col + 1, value="EA latest"),
               fill=FILL_INPUT)
        dvd = DataValidation(type="list",
                             formula1='"Manual,EA latest,EIA feed"',
                             allow_blank=False)
        ws.add_data_validation(dvd)
        dvd.add(f"{vl}{r}")
        src = f"${vl}${r}"
        r += 1
        ws.cell(row=r, column=col, value="Manual entry (kbd)").font = FONT_SMALL
        _style(ws.cell(row=r, column=col + 1, value=last_dem),
               fill=FILL_INPUT, fmt="#,##0")
        man = f"${vl}${r}"
        r += 1
        ws.cell(row=r, column=col,
                value="EA latest month (live from US Balance)").font = FONT_SMALL
        _style(ws.cell(row=r, column=col + 1,
                       value=(f'=IF(COUNT({ea_rng})=0,"",'
                              f"INDEX({ea_rng},COUNT({ea_rng})))")),
               fill=FILL_CALC, fmt="#,##0.0")
        ea = f"${vl}${r}"
        r += 1
        ws.cell(row=r, column=col,
                value="EIA petchem naphtha (Live Feeds tab)").font = FONT_SMALL
        _style(ws.cell(row=r, column=col + 1, value="='Live Feeds'!$C$10"),
               fill=FILL_CALC, fmt="#,##0.0")
        ws.cell(row=r, column=col + 2,
                value="proxy: product supplied, excl. blending"
                ).font = FONT_SMALL
        eia = f"${vl}${r}"
        r += 1
        _style(ws.cell(row=r, column=col, value="DEMAND USED"),
               fill=FILL_TOTAL, bold=True)
        _style(ws.cell(
            row=r, column=col + 1,
            value=(f'=IF({src}="Manual",{man},'
                   f'IF({src}="EA latest",IF(ISNUMBER({ea}),{ea},{man}),'
                   f"IF(ISNUMBER({eia}),{eia},{man})))")),
            fill=FILL_TOTAL, fmt="#,##0.0", bold=True)
        self.demand_ref = f"Assumptions!${vl}${r}"
        r += 2

        # ---- sensitivity grids (exact: the model is linear in the dials)
        BOX = "'Individual Refineries'"
        base_l = get_column_letter(self.c_base)
        net_expr = (f'SUMPRODUCT(({BOX}!$A$3:$A${self.box_last}="TOTAL")'
                    f"*{BOX}!${base_l}$3:${base_l}${self.box_last})")
        u0 = self.dial_refs["Utilization scaler (100% = as-is)"]
        y0 = self.dial_refs["Naphtha yield scaler (100% = as-is)"]
        _hdr(ws, r, col, "SENSITIVITY — US net naphtha (kbd)")
        ws.cell(row=r, column=col + 1,
                value="utilization scaler across, yield scaler down").font = FONT_SMALL
        r += 1
        base_row = r
        ws.cell(row=r, column=col, value="base @ 100/100").font = FONT_SMALL
        _style(ws.cell(row=r, column=col + 1,
                       value=f"={net_expr}/{u0}/{y0}"),
               fill=FILL_CALC, fmt="#,##0.0")
        base_cell = f"${vl}${base_row}"
        scalers = [0.9, 0.95, 1.0, 1.05, 1.1]
        r += 1
        for j, u in enumerate(scalers):
            _hdr(ws, r, col + 1 + j, f"util {u:.0%}")
        for yv in scalers:
            r += 1
            _style(ws.cell(row=r, column=col, value=f"yield {yv:.0%}"),
                   fill=FILL_CALC)
            for j, u in enumerate(scalers):
                cell = ws.cell(row=r, column=col + 1 + j,
                               value=f"={base_cell}*{u}*{yv}")
                _style(cell, fill=FILL_TOTAL if (u == 1 and yv == 1) else FILL_CALC,
                       fmt="#,##0")
        r += 2

        _hdr(ws, r, col, "SENSITIVITY — arbs vs freight ($/bbl)")
        ws.cell(row=r, column=col + 1,
                value="freight multiplier across; positive = arb OPEN").font = FONT_SMALL
        r += 1
        mults = [0.5, 0.75, 1.0, 1.25, 1.5]
        for j, m in enumerate(mults):
            _hdr(ws, r, col + 1 + j, f"{m:.0%} frt")
        P = self.price_refs
        arb_specs = [
            ("USGC -> Asia",
             f'({P["naphtha_asia"]}-{P["naphtha_usgc"]})',
             self.freight_refs["usgc_asia"]),
            ("USGC -> NWE",
             f'({P["naphtha_nwe"]}-{P["naphtha_usgc"]})',
             self.freight_refs["usgc_nwe"]),
            ("NWE -> USGC",
             f'({P["naphtha_usgc"]}-{P["naphtha_nwe"]})',
             self.freight_refs["nwe_usgc"]),
        ]
        from openpyxl.formatting.rule import CellIsRule as _CIR

        for label, legs, frt in arb_specs:
            r += 1
            _style(ws.cell(row=r, column=col, value=label), fill=FILL_CALC)
            for j, m in enumerate(mults):
                _style(ws.cell(row=r, column=col + 1 + j,
                               value=f"={legs}-{frt}*{m}"),
                       fill=FILL_CALC, fmt="+0.00;-0.00")
        rng = (f"{get_column_letter(col + 1)}{r - 2}:"
               f"{get_column_letter(col + len(mults))}{r}")
        ws.conditional_formatting.add(
            rng, _CIR(operator="greaterThan", formula=["0"], fill=FILL_PASS))

        ws.column_dimensions[cl].width = 40
        ws.column_dimensions[vl].width = 12
        ws.column_dimensions[tl].width = 14
        ws.column_dimensions[nl].width = 26
        ws.column_dimensions[get_column_letter(col + 4)].width = 24


    LINK_FONT = Font(size=9, color="1F5AA8", underline="single", name="Calibri")

    def _link_to_box(self, cell, rid: str) -> None:
        row = getattr(self, "box_banner_rows", {}).get(rid)
        if row:
            cell.hyperlink = f"#'Individual Refineries'!A{row}"
            cell.font = self.LINK_FONT

    # -------------------------------------------------------- Nameplate tab

    UNIT_COLS = ["CDU", "VDU", "FCC", "RCC", "COKER", "FCOKER", "DHCU",
                 "RHCU", "REF", "CCR", "ISOM", "ALKY", "NSPL", "NHT"]

    def _grid_refineries(self):
        return sorted(self.data.refineries,
                      key=lambda x: (x.padd, -x.crude_capacity_kbd, x.name))

    def _sheet_nameplate(self) -> None:
        """Stated capacity pivot: refinery x unit, live SUMIFS over the boxes so
        capacity edits there flow through."""
        ws = self.wb.create_sheet("Nameplate")
        _banner(ws, 1, "Nameplate capacity (kbd) — stated unit capacities, live from "
                       "Individual Refineries. This is what's on paper; see Effective for "
                       "what units have actually demonstrated.", 5 + len(self.UNIT_COLS))
        headers = ["refinery_id", "name", "padd", "crude kbd"] + self.UNIT_COLS
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 2, c, h)
        b_rid = f"'Individual Refineries'!$B$3:$B${self.box_last}"
        b_uid = f"'Individual Refineries'!$D$3:$D${self.box_last}"
        b_cap = f"'Individual Refineries'!$F$3:$F${self.box_last}"
        d_id = f"Data!$A$3:$A${self.data_last_row}"
        d_cap = f"Data!$G$3:$G${self.data_last_row}"
        r = 3
        for ref in self._grid_refineries():
            c = _style(ws.cell(row=r, column=1, value=ref.refinery_id))
            self._link_to_box(c, ref.refinery_id)
            _style(ws.cell(row=r, column=2, value=ref.name))
            _style(ws.cell(row=r, column=3, value=ref.padd))
            _style(ws.cell(row=r, column=4,
                           value=f"=SUMIFS({d_cap},{d_id},$A{r})"),
                   fill=FILL_CALC, fmt="#,##0")
            for j, uid in enumerate(self.UNIT_COLS):
                _style(ws.cell(
                    row=r, column=5 + j,
                    value=f'=SUMIFS({b_cap},{b_rid},$A{r},{b_uid},"{uid}")',
                ), fill=FILL_CALC, fmt="#,##0")
            r += 1
        last = r - 1
        r += 1
        _banner(ws, r, "PADD totals (kbd)", 5 + len(self.UNIT_COLS))
        for p in self.padds:
            r += 1
            _style(ws.cell(row=r, column=1, value=f"PADD {p}"), bold=True)
            _style(ws.cell(row=r, column=4,
                           value=f"=SUMIFS($D$3:$D${last},$C$3:$C${last},{p})"),
                   fill=FILL_TOTAL, fmt="0", bold=True)
            for j in range(len(self.UNIT_COLS)):
                cl = get_column_letter(5 + j)
                _style(ws.cell(row=r, column=5 + j,
                               value=f"=SUMIFS({cl}3:{cl}{last},$C$3:$C${last},{p})"),
                       fill=FILL_TOTAL, fmt="0", bold=True)
        self.nameplate_last = last
        widths = [22, 30, 6, 9] + [7] * len(self.UNIT_COLS)
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "E3"

    # -------------------------------------------------------- Effective tab

    def _sheet_effective(self) -> None:
        """Demonstrated capacity: max annual throughput 2017-2024 excl. 2020
        (RDT actuals). Same grid as Nameplate so the two compare row by row."""
        ws = self.wb.create_sheet("Effective")
        _banner(ws, 1, "Effective capacity (kbd) — max demonstrated annual throughput "
                       "2017-2024 excl. 2020 (RefineryDataTool actuals). BLUE = "
                       "desk-adjustable: edits here flow into the boxes' 'eff cap' "
                       "column. Nameplate is what's stated; this is what's proven.",
                7 + len(self.UNIT_COLS))
        headers = (["refinery_id", "name", "padd", "crude kbd"] + self.UNIT_COLS
                   + ["CDU eff yr", "CDU eff/plate"])
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 2, c, h)
        d_id = f"Data!$A$3:$A${self.data_last_row}"
        d_cap = f"Data!$G$3:$G${self.data_last_row}"
        r = 3
        for ref in self._grid_refineries():
            _style(ws.cell(row=r, column=1, value=ref.refinery_id))
            _style(ws.cell(row=r, column=2, value=ref.name))
            _style(ws.cell(row=r, column=3, value=ref.padd))
            _style(ws.cell(row=r, column=4,
                           value=f"=SUMIFS({d_cap},{d_id},$A{r})"),
                   fill=FILL_CALC, fmt="#,##0")
            for j, uid in enumerate(self.UNIT_COLS):
                _style(ws.cell(row=r, column=5 + j,
                               value=self.eff_caps.get((ref.refinery_id, uid))),
                       fill=FILL_INPUT, fmt="#,##0")
            _style(ws.cell(row=r, column=5 + len(self.UNIT_COLS),
                           value=self.eff_years.get((ref.refinery_id, "CDU"))))
            _style(ws.cell(
                row=r, column=6 + len(self.UNIT_COLS),
                value=f'=IFERROR($E{r}/Nameplate!$E{r},"")',
            ), fill=FILL_CALC, fmt="0%")
            r += 1
        last = r - 1
        r += 1
        _banner(ws, r, "PADD rollup — CDU nameplate vs effective vs running (kbd, "
                       "unit-detail refineries only)", 8)
        _hdr(ws, r + 1, 1, "PADD")
        _hdr(ws, r + 1, 2, "nameplate")
        _hdr(ws, r + 1, 3, "effective")
        _hdr(ws, r + 1, 4, "running now")
        b_pad = f"'Individual Refineries'!$C$3:$C${self.box_last}"
        b_uid = f"'Individual Refineries'!$D$3:$D${self.box_last}"
        b_cap = f"'Individual Refineries'!$F$3:$F${self.box_last}"
        b_utl = f"'Individual Refineries'!${get_column_letter(self.c_util)}$3:" \
                f"${get_column_letter(self.c_util)}${self.box_last}"
        for i, p in enumerate(self.padds):
            rr = r + 2 + i
            _style(ws.cell(row=rr, column=1, value=p), bold=True)
            _style(ws.cell(
                row=rr, column=2,
                value=f"=SUMIFS(Nameplate!$E$3:$E${self.nameplate_last},"
                      f"Nameplate!$C$3:$C${self.nameplate_last},{p})",
            ), fill=FILL_TOTAL, fmt="#,##0", bold=True)
            _style(ws.cell(row=rr, column=3,
                           value=f"=SUMIFS($E$3:$E${last},$C$3:$C${last},{p})"),
                   fill=FILL_TOTAL, fmt="#,##0", bold=True)
            _style(ws.cell(
                row=rr, column=4,
                value=(f'=SUMPRODUCT(({b_uid}="CDU")'
                       f"*({b_pad}={p})*{b_cap}*{b_utl})"),
            ), fill=FILL_TOTAL, fmt="#,##0", bold=True)
        self.eff_rollup_first = r + 2   # first PADD row of the rollup block
        # gold data bars on demonstrated CDU capacity
        from openpyxl.formatting.rule import DataBarRule

        ws.conditional_formatting.add(
            f"E3:E{last}",
            DataBarRule(start_type="num", start_value=0, end_type="max",
                        color=GOLD, showValue=True),
        )
        widths = [22, 30, 6, 9] + [7] * len(self.UNIT_COLS) + [9, 10]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "E3"


    # --------------------------------------------------------- Outages tab

    def _sheet_outages_tab(self) -> None:
        """Forward TAR calendar from the desk offline-events export: weekly
        supply-at-risk strip (anchored to TODAY(), so it stays current) over
        a filterable event table. LIVE flags recompute on open."""
        import csv as _csv
        from datetime import date as _date

        ws = self.wb.create_sheet("Outages")
        _banner(
            ws, 1,
            "CURRENT & PLANNED outages — the Snowflake live pull lands in this "
            "table (wire the query here; refreshed today via ingest-outages). "
            "Weekly strip and LIVE flags track TODAY(); live events prefill "
            "OFFLINE % in the boxes. Scenario/unplanned simulation layer: "
            "future improvement. History for forecasting: see Outage History.",
            13,
        )

        # events: live or future, next 12 months
        ev_path = DATA_DIR / "reference" / "outage_events.csv"
        events = []
        if ev_path.exists():
            as_of = self.axis[0].isoformat()
            horizon = self.axis[0].replace(year=self.axis[0].year + 1).isoformat()
            with ev_path.open() as fh:
                for e in _csv.DictReader(fh):
                    if (e["refinery_id"] and e["model_unit"]
                            and e["end"] >= as_of and e["start"] <= horizon):
                        events.append(e)
            events.sort(key=lambda e: (e["start"], e["refinery_id"]))

        # ---- weekly supply-at-risk strip (13 weeks, prorated by overlap)
        WEEKS = 13
        first_ev = 13                      # events table starts here (row)
        last_ev = first_ev + max(len(events), 1) - 1
        S, E = f"$D${first_ev}:$D${last_ev}", f"$E${first_ev}:$E${last_ev}"
        OFF = f"$I${first_ev}:$I${last_ev}"
        UNIT = f"$C${first_ev}:$C${last_ev}"
        PADD = f"$B${first_ev}:$B${last_ev}"

        _banner(ws, 3, "  CAPACITY OFFLINE BY WEEK (kbd, prorated)", 2 + WEEKS)
        _hdr(ws, 4, 1, "week starting")
        for k in range(WEEKS):
            cell = ws.cell(row=4, column=2 + k,
                           value=f"=TODAY()-WEEKDAY(TODAY(),3)+{k * 7}")
            _style(cell, fill=FILL_HDR, fmt="m/d")
            cell.font = FONT_HDR

        def overlap(ws_ref, we_ref):
            min_e = f"(({E}+{we_ref}-ABS({E}-{we_ref}))/2)"
            max_s = f"(({S}+{ws_ref}+ABS({S}-{ws_ref}))/2)"
            od = f"({min_e}-{max_s}+1)"
            return f"(({od}+ABS({od}))/2)"

        strips = [
            ("US — CDU offline", f'({UNIT}="CDU")'),
            ("PADD 3 — CDU offline", f'({UNIT}="CDU")*({PADD}=3)'),
            ("US — reformer offline (REF+CCR)",
             f'(({UNIT}="REF")+({UNIT}="CCR"))'),
        ]
        for i, (label, flt) in enumerate(strips):
            r = 5 + i
            _style(ws.cell(row=r, column=1, value=label), fill=FILL_CALC, bold=True)
            for k in range(WEEKS):
                col = get_column_letter(2 + k)
                ws_ref, we_ref = f"{col}$4", f"({col}$4+6)"
                _style(ws.cell(
                    row=r, column=2 + k,
                    value=(f"=SUMPRODUCT({flt}*{OFF}"
                           f"*{overlap(ws_ref, we_ref)}/7)"),
                ), fill=FILL_CALC, fmt="#,##0")

        from openpyxl.chart import BarChart, Reference, Series

        ch = BarChart()
        ch.type = "col"
        ch.title = "CDU capacity offline by week (kbd)"
        ch.height, ch.width = 6.5, 22
        ch.legend = None
        ser = Series(Reference(ws, min_col=2, max_col=1 + WEEKS, min_row=5, max_row=5))
        ser.graphicalProperties.solidFill = "B03A2E"
        ch.series.append(ser)
        ch.set_categories(Reference(ws, min_col=2, max_col=1 + WEEKS, min_row=4, max_row=4))
        _axes(ch, "week starting", "kbd offline", rotate=True)
        ws.add_chart(ch, f"P3")

        # ---- events table
        _banner(ws, first_ev - 2, "  EVENTS — live + next 12 months", 13)
        headers = ["refinery", "padd", "unit", "start", "end", "days", "type",
                   "live?", "offline kbd", "% of unit", "confirmed", "cause",
                   "plant"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, first_ev - 1, c, h)
        padd_of = {r.refinery_id: r.padd for r in self.data.refineries}
        for j, e in enumerate(events):
            r = first_ev + j
            cell = _style(ws.cell(row=r, column=1, value=e["refinery_id"]))
            self._link_to_box(cell, e["refinery_id"])
            _style(ws.cell(row=r, column=2, value=padd_of.get(e["refinery_id"])))
            _style(ws.cell(row=r, column=3, value=e["model_unit"]))
            _style(ws.cell(row=r, column=4,
                           value=_date.fromisoformat(e["start"])), fmt="yyyy-mm-dd")
            _style(ws.cell(row=r, column=5,
                           value=_date.fromisoformat(e["end"])), fmt="yyyy-mm-dd")
            _style(ws.cell(row=r, column=6, value=f"=$E{r}-$D{r}"), fmt="0")
            _style(ws.cell(row=r, column=7, value=e["event_type"]))
            _style(ws.cell(
                row=r, column=8,
                value=f'=IF(AND($D{r}<=TODAY(),$E{r}>=TODAY()),"LIVE","")',
            ))
            off = float(e["capacity_offline"] or 0)
            cap = float(e["unit_capacity"] or 0)
            _style(ws.cell(row=r, column=9, value=off / 1000), fmt="#,##0.0")
            _style(ws.cell(row=r, column=10,
                           value=off / cap if cap else None), fmt="0%")
            _style(ws.cell(row=r, column=11, value=e["confirmation"]))
            _style(ws.cell(row=r, column=12, value=e["cause"]))
            _style(ws.cell(row=r, column=13, value=e["plant_name"]))
        last = first_ev + len(events) - 1
        ws.auto_filter.ref = f"A{first_ev - 1}:M{last}"
        ws.conditional_formatting.add(
            f"H{first_ev}:H{last}",
            CellIsRule(operator="equal", formula=['"LIVE"'], fill=FILL_FAIL))
        widths = [24, 6, 7, 11, 11, 6, 10, 7, 11, 9, 11, 22, 30]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = f"A{first_ev}"


    # -------------------------------------------------- Outage History tab

    def _sheet_outage_history(self) -> None:
        """Historical outages (2023 -> now): monthly offline series,
        turnaround seasonality, planned-vs-unplanned stats and the full
        filterable event history - the base for forecasting planned and
        unplanned outages."""
        import csv as _csv
        from datetime import date as _date

        ws = self.wb.create_sheet("Outage History")
        _banner(
            ws, 1,
            "OUTAGE HISTORY (2023 -> today) — what normal looks like: monthly "
            "offline capacity, TAR seasonality and unplanned frequency, to "
            "benchmark forecasts of planned and unplanned outages.", 13)

        ev_path = DATA_DIR / "reference" / "outage_events.csv"
        as_of = self.axis[0]
        past = []
        if ev_path.exists():
            with ev_path.open() as fh:
                for e in _csv.DictReader(fh):
                    if e["refinery_id"] and e["model_unit"] and \
                            e["start"] < as_of.isoformat():
                        past.append(e)

        # monthly offline series (prorated), CDU and reformers
        def month_axis():
            y, m = 2023, 1
            while (y, m) <= (as_of.year, as_of.month):
                yield y, m
                y, m = (y + 1, 1) if m == 12 else (y, m + 1)

        import calendar
        series = {"CDU": [], "REFORMER": []}
        months = list(month_axis())
        for y, m in months:
            days = calendar.monthrange(y, m)[1]
            m0, m1 = _date(y, m, 1), _date(y, m, days)
            tot = {"CDU": 0.0, "REFORMER": 0.0}
            for e in past:
                grp = ("CDU" if e["model_unit"] == "CDU" else
                       "REFORMER" if e["model_unit"] in ("REF", "CCR") else None)
                if grp is None:
                    continue
                s0 = _date.fromisoformat(e["start"])
                s1 = _date.fromisoformat(e["end"])
                ov = (min(s1, m1) - max(s0, m0)).days + 1
                if ov > 0:
                    tot[grp] += float(e["capacity_offline"] or 0) * ov / days
            for k in series:
                series[k].append(tot[k] / 1000.0)

        _banner(ws, 3, "  CAPACITY OFFLINE BY MONTH (kbd, prorated history)",
                2 + len(months))
        _hdr(ws, 4, 1, "month")
        for j, (y, m) in enumerate(months):
            cell = ws.cell(row=4, column=2 + j, value=_date(y, m, 1))
            _style(cell, fill=FILL_HDR, fmt="mmm-yy")
            cell.font = FONT_HDR
        for i, (label, vals) in enumerate(
                [("US CDU offline", series["CDU"]),
                 ("US reformer offline", series["REFORMER"])]):
            _style(ws.cell(row=5 + i, column=1, value=label),
                   fill=FILL_CALC, bold=True)
            for j, v in enumerate(vals):
                _style(ws.cell(row=5 + i, column=2 + j, value=round(v, 1)),
                       fill=FILL_CALC, fmt="#,##0")

        # TAR seasonality: average CDU offline by calendar month
        _banner(ws, 8, "  TAR SEASONALITY — avg CDU kbd offline by calendar month", 13)
        for m in range(1, 13):
            vals = [series["CDU"][j] for j, (y, mm) in enumerate(months) if mm == m]
            _hdr(ws, 9, 1 + m, _date(2000, m, 1).strftime("%b"))
            _style(ws.cell(row=10, column=1 + m,
                           value=round(sum(vals) / len(vals), 1) if vals else None),
                   fill=FILL_CALC, fmt="#,##0")
        _style(ws.cell(row=10, column=1, value="avg kbd offline"),
               fill=FILL_CALC, bold=True)

        # planned vs unplanned stats
        _banner(ws, 12, "  PLANNED vs UNPLANNED (naphtha-relevant events, 2023 ->)", 13)
        for c, h in enumerate(["type", "events", "avg days", "avg kbd offline",
                               "offline kbd-days / yr"], start=1):
            _hdr(ws, 13, c, h)
        yrs = max((as_of - _date(2023, 1, 1)).days / 365.25, 1e-9)
        for i, typ in enumerate(["Planned", "Unplanned"]):
            evs = [e for e in past if e["event_type"] == typ]
            durs = [float(e["duration_days"] or 0) for e in evs]
            offs = [float(e["capacity_offline"] or 0) / 1000 for e in evs]
            kbd_days = sum(o * d for o, d in zip(offs, durs))
            r = 14 + i
            _style(ws.cell(row=r, column=1, value=typ), bold=True)
            _style(ws.cell(row=r, column=2, value=len(evs)), fmt="#,##0")
            _style(ws.cell(row=r, column=3,
                           value=round(sum(durs) / max(len(durs), 1), 1)), fmt="0.0")
            _style(ws.cell(row=r, column=4,
                           value=round(sum(offs) / max(len(offs), 1), 1)), fmt="0.0")
            _style(ws.cell(row=r, column=5, value=round(kbd_days / yrs)), fmt="#,##0")

        from openpyxl.chart import BarChart, LineChart, Reference, Series

        ch = LineChart()
        ch.title = "US CDU capacity offline by month (kbd)"
        ch.height, ch.width = 7, 24
        ch.legend = None
        ser = Series(Reference(ws, min_col=2, max_col=1 + len(months),
                               min_row=5, max_row=5))
        ser.graphicalProperties.line.solidFill = "B03A2E"
        ch.series.append(ser)
        ch.set_categories(Reference(ws, min_col=2, max_col=1 + len(months),
                                    min_row=4, max_row=4))
        _axes(ch, "month", "kbd offline", rotate=True, skip=3)
        ws.add_chart(ch, "P3")

        # event history detail
        first = 19
        _banner(ws, first - 2, "  EVENT HISTORY — filterable detail", 13)
        headers = ["refinery", "padd", "unit", "start", "end", "days", "type",
                   "offline kbd", "% of unit", "confirmed", "cause", "plant"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, first - 1, c, h)
        padd_of = {r.refinery_id: r.padd for r in self.data.refineries}
        past.sort(key=lambda e: e["start"], reverse=True)
        for j, e in enumerate(past):
            r = first + j
            cell = _style(ws.cell(row=r, column=1, value=e["refinery_id"]))
            self._link_to_box(cell, e["refinery_id"])
            _style(ws.cell(row=r, column=2, value=padd_of.get(e["refinery_id"])))
            _style(ws.cell(row=r, column=3, value=e["model_unit"]))
            _style(ws.cell(row=r, column=4,
                           value=_date.fromisoformat(e["start"])), fmt="yyyy-mm-dd")
            _style(ws.cell(row=r, column=5,
                           value=_date.fromisoformat(e["end"])), fmt="yyyy-mm-dd")
            _style(ws.cell(row=r, column=6,
                           value=float(e["duration_days"] or 0)), fmt="0")
            _style(ws.cell(row=r, column=7, value=e["event_type"]))
            off = float(e["capacity_offline"] or 0)
            cap = float(e["unit_capacity"] or 0)
            _style(ws.cell(row=r, column=8, value=off / 1000), fmt="#,##0.0")
            _style(ws.cell(row=r, column=9,
                           value=off / cap if cap else None), fmt="0%")
            _style(ws.cell(row=r, column=10, value=e["confirmation"]))
            _style(ws.cell(row=r, column=11, value=e["cause"]))
            _style(ws.cell(row=r, column=12, value=e["plant_name"]))
        ws.auto_filter.ref = f"A{first - 1}:L{first + len(past) - 1}"
        widths = [24, 6, 7, 11, 11, 6, 10, 11, 9, 11, 22, 30]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = f"A{first}"

    # ------------------------------------------------------ Kpler Flows tab

    TRADE_ROWS = 400
    KPLER_GRADES = ["Light Naphtha", "Virgin Naphtha", "LVN", "Heavy Naphtha",
                    "Naphtha Diluent", "Other"]
    KPLER_STATUS = ["delivered", "scheduled", "loading", "in transit"]

    def _sheet_kpler(self) -> None:
        """Kpler ship-tracking landing zones: TRADES (the core dataset -
        locations, products, direction, intra-country vs intra-region),
        plus FIXTURES and FLOWS tables. Live aggregations by month, grade
        and status feed the US Balance tab."""
        ws = self.wb.create_sheet("Kpler Flows")
        _banner(
            ws, 1,
            "KPLER SHIP TRACKING — paste exports here. TRADES is the core set "
            "(break down by location, product, direction, intra-country vs "
            "intra-region); FIXTURES and FLOWS below. CAVEAT: Kpler misses "
            "up-river legs in some regions (LatAm), so country-to-country "
            "shipments can be absent from supply — less of a concern for the US.",
            16,
        )

        # ---------------- TRADES
        _banner(ws, 3, "  TRADES — the core dataset", 14)
        headers = ["date", "grade", "direction", "origin country", "origin zone",
                   "dest country", "dest zone", "scope", "volume kbd", "status",
                   "vessel", "charterer", "source", "notes"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 4, c, h)
        t0, t1 = 5, 4 + self.TRADE_ROWS
        for r in range(t0, t1 + 1):
            for c in range(1, 15):
                _style(ws.cell(row=r, column=c), fill=FILL_INPUT,
                       fmt="yyyy-mm-dd" if c == 1 else ("0.0" if c == 9 else None))
        for col, options in (("B", ",".join(self.KPLER_GRADES)),
                             ("C", "import,export,intra-US"),
                             ("H", "international,intra-region,intra-country"),
                             ("J", ",".join(self.KPLER_STATUS))):
            dv = DataValidation(type="list", formula1=f'"{options}"',
                                allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col}{t0}:{col}{t1}")

        dcol, gcol = f"$A${t0}:$A${t1}", f"$B${t0}:$B${t1}"
        ccol, vcol = f"$C${t0}:$C${t1}", f"$I${t0}:$I${t1}"
        scol = f"$J${t0}:$J${t1}"

        # monthly imports/exports (last 12 months, live)
        P = 16
        _banner(ws, 3, "", 1)  # no-op to keep spacing simple
        _hdr(ws, 4, P, "month")
        _hdr(ws, 4, P + 1, "imports kbd")
        _hdr(ws, 4, P + 2, "exports kbd")
        _hdr(ws, 4, P + 3, "net kbd")
        for k in range(12):
            r = 5 + k
            mcell = f"${get_column_letter(P)}{r}"
            _style(ws.cell(row=r, column=P, value=f"=EOMONTH(TODAY(),{k - 12})+1"),
                   fill=FILL_CALC, fmt="mmm-yy")
            for j, d in enumerate(("import", "export")):
                _style(ws.cell(
                    row=r, column=P + 1 + j,
                    value=(f'=SUMIFS({vcol},{ccol},"{d}",{dcol},">="&{mcell},'
                           f'{dcol},"<"&EOMONTH({mcell},0)+1)'),
                ), fill=FILL_CALC, fmt="0.0")
            _style(ws.cell(row=r, column=P + 3,
                           value=f"={get_column_letter(P + 1)}{r}"
                                 f"-{get_column_letter(P + 2)}{r}"),
                   fill=FILL_CALC, fmt="+0.0;-0.0")
        self.kpler_cur_imports = f"'Kpler Flows'!${get_column_letter(P + 1)}$16"
        self.kpler_cur_exports = f"'Kpler Flows'!${get_column_letter(P + 2)}$16"

        # by grade (trailing 90 days)
        _hdr(ws, 19, P, "grade (trailing 90d)")
        _hdr(ws, 19, P + 1, "imports kbd")
        _hdr(ws, 19, P + 2, "exports kbd")
        for i, g in enumerate(self.KPLER_GRADES):
            r = 20 + i
            _style(ws.cell(row=r, column=P, value=g), fill=FILL_CALC)
            for j, d in enumerate(("import", "export")):
                _style(ws.cell(
                    row=r, column=P + 1 + j,
                    value=(f'=SUMIFS({vcol},{gcol},"{g}",{ccol},"{d}",'
                           f'{dcol},">="&(TODAY()-90))'),
                ), fill=FILL_CALC, fmt="0.0")

        # by status (all rows)
        _hdr(ws, 28, P, "status")
        _hdr(ws, 28, P + 1, "total kbd")
        for i, st in enumerate(self.KPLER_STATUS):
            r = 29 + i
            _style(ws.cell(row=r, column=P, value=st), fill=FILL_CALC)
            _style(ws.cell(row=r, column=P + 1,
                           value=f'=SUMIFS({vcol},{scol},"{st}")'),
                   fill=FILL_CALC, fmt="0.0")

        # ---------------- FIXTURES
        f0 = t1 + 4
        _banner(ws, f0 - 2, "  FIXTURES", 11)
        fx_headers = ["fixture date", "laycan start", "laycan end", "vessel",
                      "grade", "qty kbd", "origin", "destination", "status",
                      "rate", "notes"]
        for c, h in enumerate(fx_headers, start=1):
            _hdr(ws, f0 - 1, c, h)
        for r in range(f0, f0 + 100):
            for c in range(1, 12):
                _style(ws.cell(row=r, column=c), fill=FILL_INPUT,
                       fmt="yyyy-mm-dd" if c <= 3 else ("0.0" if c == 6 else None))
        dv = DataValidation(type="list", formula1=f'"{",".join(self.KPLER_STATUS)}"',
                            allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"I{f0}:I{f0 + 99}")

        # ---------------- FLOWS (aggregated bilateral)
        w0 = f0 + 104
        _banner(ws, w0 - 2, "  FLOWS — aggregated origin -> destination", 7)
        fl_headers = ["month", "origin zone", "dest zone", "grade", "kbd",
                      "source", "notes"]
        for c, h in enumerate(fl_headers, start=1):
            _hdr(ws, w0 - 1, c, h)
        for r in range(w0, w0 + 100):
            for c in range(1, 8):
                _style(ws.cell(row=r, column=c), fill=FILL_INPUT,
                       fmt="mmm-yy" if c == 1 else ("0.0" if c == 5 else None))

        widths = [11, 14, 9, 13, 12, 13, 12, 13, 10, 10, 15, 12, 10, 20, 3,
                  10, 11, 11, 9]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A5"

    # ------------------------------------------------------- US Balance tab

    def _sheet_usbalance(self) -> None:
        """US total naphtha balance: EA monthly actuals (2023 -> present)
        plus a live tie-in of the model's net naphtha and Kpler flows."""
        import csv as _csv
        from datetime import date as _date

        ws = self.wb.create_sheet("US Balance")
        _banner(
            ws, 1,
            "US NAPHTHA BALANCE — EA monthly actuals below; live tie-in on top "
            "combines the model's net naphtha with Kpler flows to nowcast the "
            "balance.", 12)

        # live tie-in block
        base_l = get_column_letter(self.c_base)
        b_typ = f"'Individual Refineries'!$A$3:$A${self.box_last}"
        b_net = (f"'Individual Refineries'!${base_l}$3:"
                 f"${base_l}${self.box_last}")
        _banner(ws, 3, "  LIVE TIE-IN (kbd)", 6)
        ties = [
            ("Model net naphtha supply (boxes, now)",
             f'=SUMPRODUCT(({b_typ}="TOTAL")*{b_net})', "#,##0.0", FILL_CALC),
            ("Kpler imports (current month)",
             f"={self.kpler_cur_imports}", "0.0", FILL_CALC),
            ("Kpler exports (current month)",
             f"={self.kpler_cur_exports}", "0.0", FILL_CALC),
            ("Demand (kbd) — source picked on the Assumptions DEMAND block",
             f"={self.demand_ref}", "#,##0.0", FILL_CALC),
            ("IMPLIED BALANCE (supply + imports - exports - demand)",
             "=$C$4+$C$5-$C$6-$C$7", "+#,##0.0;-#,##0.0", FILL_TOTAL),
        ]
        for i, (label, formula, fmt, fill) in enumerate(ties):
            r = 4 + i
            _style(ws.cell(row=r, column=1, value=label),
                   fill=FILL_CALC, bold=(i == 4))
            _style(ws.cell(row=r, column=3, value=formula), fill=fill,
                   fmt=fmt, bold=(i == 4))

        # EA monthly actuals
        bal_path = DATA_DIR / "reference" / "us_naphtha_balance_monthly.csv"
        flows = ["REFGROUT", "TOTIMPSB", "TOTEXPSB", "TOTDEMO", "BLENDING",
                 "STOCKCH", "CLOSTLV", "BALANCE"]
        labels = ["refinery gross out", "imports", "exports", "demand",
                  "blending", "stock change", "closing stocks", "balance"]
        table: dict[str, dict[str, float]] = {}
        if bal_path.exists():
            with bal_path.open() as fh:
                for row in _csv.DictReader(fh):
                    table.setdefault(row["month"], {})[row["flow"]] = float(
                        row["kbd"] or 0)
        months = sorted(table)
        first = 12
        _banner(ws, first - 2, "  EA MONTHLY ACTUALS (kbd)", 2 + len(flows))
        _hdr(ws, first - 1, 1, "month")
        for c, lab in enumerate(labels, start=2):
            _hdr(ws, first - 1, c, lab)
        for j, m in enumerate(months):
            r = first + j
            _style(ws.cell(row=r, column=1,
                           value=_date.fromisoformat(m + "-01")), fmt="mmm-yy")
            for c, fl in enumerate(flows, start=2):
                _style(ws.cell(row=r, column=c, value=table[m].get(fl)),
                       fmt="#,##0" if fl == "CLOSTLV" else "0.0")
        last = first + len(months) - 1
        self.usbal_first, self.usbal_last = first, last

        from openpyxl.chart import LineChart, Reference, Series

        ch = LineChart()
        ch.title = "US naphtha: imports, exports, balance (kbd, monthly)"
        ch.height, ch.width = 8, 22
        for col, lab, color in ((3, "imports", "1F6FB2"),
                                (4, "exports", "B03A2E"),
                                (9, "balance", GOLD)):
            ser = Series(Reference(ws, min_col=col, min_row=first, max_row=last),
                         title=lab)
            ser.graphicalProperties.line.solidFill = color
            ch.series.append(ser)
        ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
        _axes(ch, "month", "kbd", rotate=True, skip=3)
        ws.add_chart(ch, "F3")
        widths = [34, 14, 12, 12, 12, 12, 12, 13, 12]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = f"A{first}"

    # ------------------------------------------------------- Live Feeds tab

    # EIA series verified against eia.gov (dnav): label, seriesid, unit
    EIA_FEED_SERIES = [
        ("US petchem naphtha demand — product supplied (monthly)",
         "PET.MNFUPUS2.M", "kbd"),
        ("US gross inputs to refineries (weekly)", "PET.WGIRIUS2.W", "kbd"),
        ("US refinery crude runs (weekly)", "PET.WCRRIUS2.W", "kbd"),
        ("US operable refining capacity (weekly)", "PET.WOCLEUS2.W", "kbd"),
        ("US refinery utilization (weekly)", "PET.WPULEUS3.W", "%"),
    ]

    def _sheet_livefeeds(self) -> None:
        """API pulls that live IN the workbook: WEBSERVICE()/FILTERXML()
        formulas hit the EIA API (free key) and the desk's IIR query so the
        numbers refresh whenever the weekly data drops — no Python needed.
        Keys/tokens are typed by the user and are never written by the build.
        The Assumptions DEMAND block reads C10 (EIA naphtha petchem demand)."""
        ws = self.wb.create_sheet("Live Feeds")
        _banner(ws, 1,
                "LIVE DATA FEEDS — API pulls inside the workbook: paste a key,"
                " hit Data > Refresh All (Ctrl+Alt+F9) when the weekly data"
                " drops.", 9)
        notes = [
            "WEBSERVICE/FILTERXML refresh in Windows desktop Excel. Offline / "
            "LibreOffice / Mac they error quietly and the model keeps the "
            "last saved numbers.",
            "Keys & tokens are typed HERE by the user - the repo build never "
            "writes them. Don't commit a saved copy that contains them.",
            "Python fallback (refreshes the CSVs the build ingests):  "
            "python -m naphtha_model.cli pull-eia --api-key ...   |   "
            "python -m naphtha_model.cli pull-iir --url ... --token ...",
        ]
        for i, n in enumerate(notes):
            ws.cell(row=3 + i, column=1, value=n).font = FONT_SMALL

        # ---- EIA block (rows 7-16; Assumptions references $C$10)
        _hdr(ws, 7, 1, "EIA — US OFFICIAL WEEKLY/MONTHLY DATA "
                        "(free key: eia.gov/opendata/register.php)")
        ws.cell(row=8, column=1, value="EIA API key").font = FONT_SMALL
        _style(ws.cell(row=8, column=3, value=None), fill=FILL_INPUT)
        ws.cell(row=8, column=4, value="xpath value / period").font = FONT_SMALL
        _style(ws.cell(row=8, column=5, value="//row[1]/value"),
               fill=FILL_INPUT)
        _style(ws.cell(row=8, column=6, value="//row[1]/period"),
               fill=FILL_INPUT)
        for c, h in enumerate(["series", "EIA series ID", "latest", "period",
                               "unit", "request URL (auto-built)"], start=1):
            _hdr(ws, 9, c, h)
        for i, (label, sid, unit) in enumerate(self.EIA_FEED_SERIES):
            r = 10 + i
            ws.cell(row=r, column=1, value=label).font = FONT_SMALL
            _style(ws.cell(row=r, column=2, value=sid), fill=FILL_INPUT)
            _style(ws.cell(
                row=r, column=3,
                value=(f'=IF($C$8="","← key",IFERROR(NUMBERVALUE('
                       f"FILTERXML(WEBSERVICE($F{r}),$E$8)),"
                       f'"refresh in Excel"))')),
                fill=FILL_CALC, fmt="#,##0.0")
            _style(ws.cell(
                row=r, column=4,
                value=(f'=IF($C$8="","",IFERROR(FILTERXML('
                       f'WEBSERVICE($F{r}),$F$8),""))')),
                fill=FILL_CALC)
            _style(ws.cell(row=r, column=5, value=unit), fill=FILL_CALC)
            ws.cell(
                row=r, column=6,
                value=('="https://api.eia.gov/v2/seriesid/"&$B' + str(r) +
                       '&"?api_key="&$C$8&"&out=xml'
                       '&sort[0][column]=period&sort[0][direction]=desc'
                       '&length=2"')).font = FONT_SMALL
        ws.cell(row=16, column=1, value=(
            "MNFUPUS2 is the petchem-feedstock leg of naphtha demand "
            "(excludes gasoline-blending pull). EA stays the primary total-"
            "demand source; EIA is the accuracy-anchored fallback."
        )).font = FONT_SMALL

        # ---- IIR block (rows 18-25) — WIRED (endpoint recovered from the
        # desk workbook's Power Query; POST-only, so no WEBSERVICE)
        _hdr(ws, 18, 1, "IIR — OFFLINE EVENTS, WIRED "
                        "(Bearer token, expires every 30 days)")
        ws.cell(row=19, column=1,
                value="IIR token (CLI reads it from gitignored secrets.yaml)"
                ).font = FONT_SMALL
        _style(ws.cell(row=19, column=3, value=None), fill=FILL_INPUT)
        ws.cell(row=20, column=1,
                value="query (US refining, all statuses)").font = FONT_SMALL
        ws.cell(
            row=20, column=3,
            value="https://api.industrialinfo.com/idb/v2.6/offlineevents/"
                  "summary?eventStatusDesc=Ongoing&eventStatusDesc=Future"
                  "&eventStatusDesc=Past&industryCodeDesc=Petroleum+Refining"
                  "+%28HPI%29&physicalAddressCountryName=U.S.A."
        ).font = FONT_SMALL
        iir_notes = [
            "The API is POST-only, so WEBSERVICE (GET) can't call it. Two "
            "refresh paths: Power Query (Data > Get Data > From Web > "
            "Advanced: URL above + Authorization header 'Bearer <token>', "
            "Content-Type application/json) or the CLI below.",
            "CLI: python -m naphtha_model.cli pull-iir  ->  pulls every US "
            "refining event, refreshes outage_events.csv + "
            "current_outages.csv (as-of today), then rebuild the workbook "
            "to light up the Outages tabs.",
            "Verified live 2026-07-14: 7,466 US records -> 5,512 events "
            "2023+, 129/137 plants matched, current outages refreshed "
            "(picked up the Valero Corpus Christi CDU+VDU TAR).",
            "OEV fields mapped: offlineEventKey, eventType, eventStart/"
            "EndDate, plantName, unitTypeDesc, offlineCapacity.unitCapacity"
            " / .capacityOffline, eventCause.",
            "units/summary and units/detail (capacities, shutdown dates, "
            "future capacity) live on the same API for the capacity walk.",
        ]
        for i, n in enumerate(iir_notes):
            ws.cell(row=21 + i, column=1, value=n).font = FONT_SMALL

        # ---- EA block (rows 27-31) — WIRED (OilX REST API)
        _hdr(ws, 27, 1, "EA — ENERGY ASPECTS, WIRED (OilX REST API; "
                        "UUID key in the api_key query param)")
        ws.cell(row=28, column=1,
                value="EA api key (CLI reads secrets.yaml)").font = FONT_SMALL
        _style(ws.cell(row=28, column=3, value=None), fill=FILL_INPUT)
        ws.cell(row=29, column=1, value="request URL").font = FONT_SMALL
        ws.cell(
            row=29, column=3,
            value=('=IF($C$28="","← key","https://api.energyaspects.com/oilx'
                   '/v2/balances/country/pop?api_key="&$C$28&"&product='
                   'NAPHTHA&country=US&marginal_range=1900-01-01,")')
        ).font = FONT_SMALL
        ws.cell(row=30, column=1,
                value="raw response (JSON — eyeball only)").font = FONT_SMALL
        _style(ws.cell(
            row=30, column=3,
            value='=IF(LEFT($C$29,4)="http",WEBSERVICE($C$29),"")'),
            fill=FILL_CALC)
        ws.cell(row=31, column=1, value=(
            "Responses are JSON (FILTERXML can't parse them), so the numbers "
            "land via CLI:  python -m naphtha_model.cli pull-ea  ->  rewrites "
            "the US Balance monthly table's CSV (verified: reproduces the "
            "manual EA export exactly, plus OilX nowcast months). The "
            "Assumptions DEMAND 'EA latest' option reads it live."
        )).font = FONT_SMALL

        # ---- Kpler block (rows 33-36) — key pending
        _hdr(ws, 33, 1, "KPLER — SHIP TRACKING (key pending)")
        ws.cell(row=34, column=1,
                value="Kpler key (base64, Basic auth)").font = FONT_SMALL
        _style(ws.cell(row=34, column=3, value=None), fill=FILL_INPUT)
        kpler_notes = [
            "The key supplied 2026-07-14 fails auth (401): two characters "
            "arrived corrupted and were repaired, but it still doesn't "
            "authenticate - likely more transcription damage. Re-copy the "
            "exact string from the Kpler console into secrets.yaml.",
            "Once valid: python -m naphtha_model.cli pull-kpler --endpoint "
            "trades --params products=naphtha...  ->  data/raw/, feeding the "
            "Kpler Flows tab (trades / fixtures / flows by grade & status).",
        ]
        for i, n in enumerate(kpler_notes):
            ws.cell(row=35 + i, column=1, value=n).font = FONT_SMALL

        for c, w in [(1, 58), (2, 18), (3, 16), (4, 12), (5, 8), (6, 100)]:
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A7"

    # ------------------------------------------------------- CrudeSlate tab

    def _sheet_crudeslate(self) -> None:
        import csv as _csv

        ws = self.wb.create_sheet("CrudeSlate")
        _banner(ws, 1, "Crude slate — actual crude diet per refinery (REM, latest "
                       "year) and 2024 purchased supplementary feedstocks. Informs "
                       "the naphtha-yield discussion (light vs heavy slate).", 8)
        headers = ["refinery_id", "padd", "crude_stream", "source_country",
                   "year", "slate %"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 2, c, h)
        padd_of = {x.refinery_id: x.padd for x in self.data.refineries}
        slate_path = DATA_DIR / "reference" / "crude_slate.csv"
        r = 3
        if slate_path.exists():
            with slate_path.open() as fh:
                rows = [row for row in _csv.DictReader(fh) if row["refinery_id"]]
            latest: dict[str, str] = {}
            for row in rows:
                rid = row["refinery_id"]
                latest[rid] = max(latest.get(rid, ""), row["year"])
            rows = [row for row in rows if row["year"] == latest[row["refinery_id"]]]
            rows.sort(key=lambda x: (padd_of.get(x["refinery_id"], 9),
                                     x["refinery_id"],
                                     -float(x["slate_pct"] or 0)))
            for row in rows:
                _style(ws.cell(row=r, column=1, value=row["refinery_id"]))
                _style(ws.cell(row=r, column=2, value=padd_of.get(row["refinery_id"])))
                _style(ws.cell(row=r, column=3, value=row["crude_stream"]))
                _style(ws.cell(row=r, column=4, value=row["source_country"]))
                _style(ws.cell(row=r, column=5, value=int(row["year"])))
                _style(ws.cell(row=r, column=6,
                               value=float(row["slate_pct"]) / 100.0), fmt="0.0%")
                r += 1

        r += 1
        _banner(ws, r, "Purchased supplementary feedstocks, 2024 — naphtha/reformate "
                       "purchases mark merchant naphtha BUYERS", 8)
        r += 1
        for c, h in enumerate(["refinery_id", "padd", "feedstock", "to unit",
                               "kbd", "API", "type", "naphtha buyer?"], start=1):
            _hdr(ws, r, c, h)
        fs_path = DATA_DIR / "reference" / "feedstock_slate.csv"
        if fs_path.exists():
            with fs_path.open() as fh:
                feeds = [row for row in _csv.DictReader(fh)
                         if row["refinery_id"] and row["year"] == "2024"
                         and float(row["kbd"]) > 0]
            feeds.sort(key=lambda x: (padd_of.get(x["refinery_id"], 9), -float(x["kbd"])))
            for row in feeds:
                r += 1
                buyer = row["feedstock"].lower() in ("naphtha", "reformate")
                _style(ws.cell(row=r, column=1, value=row["refinery_id"]))
                _style(ws.cell(row=r, column=2, value=padd_of.get(row["refinery_id"])))
                _style(ws.cell(row=r, column=3, value=row["feedstock"]))
                _style(ws.cell(row=r, column=4, value=row["to_unit"]))
                _style(ws.cell(row=r, column=5, value=float(row["kbd"])), fmt="0.0")
                _style(ws.cell(row=r, column=6,
                               value=float(row["api"]) if row["api"] else None), fmt="0.0")
                _style(ws.cell(row=r, column=7, value=row["api_type"]))
                cell = ws.cell(row=r, column=8, value="BUYER" if buyer else "")
                _style(cell, fill=FILL_TOTAL if buyer else None, bold=buyer)
        widths = [24, 6, 30, 13, 8, 7, 8, 13]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A3"

    # -------------------------------------------------------- BlendEcon tab

    def _sheet_blendecon(self) -> None:
        ws = self.wb.create_sheet("BlendEcon")
        _banner(ws, 1, "Blending economics — FRAMEWORK: fill the blue price / freight "
                       "/ scenario inputs on the Assumptions tab and every number "
                       "here goes live.", 8)
        P, F, B = self.price_refs, self.freight_refs, self.blend_refs

        _banner(ws, 3, "Blend value spreads ($/bbl)", 4)
        spreads = [
            ("Gasoline - HVN (naphtha pull into mogas when positive)",
             f'={P["gasoline"]}-{P["naphtha_usgc"]}'),
            ("Gasoline - LVN",
             f'={P["gasoline"]}-{P["naphtha_lvn"]}'),
            ("Reformate - HVN (reforming uplift)",
             f'={P["reformate"]}-{P["naphtha_usgc"]}'),
            ("Naphtha USGC - WTI (naphtha crack)",
             f'={P["naphtha_usgc"]}-{P["wti"]}'),
        ]
        r = 4
        for label, formula in spreads:
            ws.cell(row=r, column=1, value=label).font = FONT_SMALL
            _style(ws.cell(row=r, column=4, value=formula), fill=FILL_CALC, fmt="0.00")
            r += 1

        r += 1
        _banner(ws, r, "Arbitrage netbacks ($/bbl) — positive = arb OPEN", 4)
        arbs = [
            ("USGC -> Asia",
             f'={P["naphtha_asia"]}-{P["naphtha_usgc"]}'
             f'-{F["usgc_asia"]}'),
            ("USGC -> NWE",
             f'={P["naphtha_nwe"]}-{P["naphtha_usgc"]}'
             f'-{F["usgc_nwe"]}'),
            ("NWE -> USGC (imports)",
             f'={P["naphtha_usgc"]}-{P["naphtha_nwe"]}'
             f'-{F["nwe_usgc"]}'),
        ]
        for label, formula in arbs:
            r += 1
            ws.cell(row=r, column=1, value=label).font = FONT_SMALL
            _style(ws.cell(row=r, column=4, value=formula), fill=FILL_CALC, fmt="+0.00;-0.00")
            arb_cell = get_column_letter(4) + str(r)
            _style(ws.cell(row=r, column=5,
                           value=f'=IF({arb_cell}>0,"OPEN","shut")'), fill=FILL_CALC)

        r += 2
        _banner(ws, r, "Slate scenarios by PADD (kbd) — max-light vs max-heavy crude", 8)
        r += 1
        for c, h in enumerate(["PADD", "crude runs", "net naphtha (base)",
                               "max-light delta", "max-heavy delta",
                               "$k/day at stake (light, vs mogas)"], start=1):
            _hdr(ws, r, c, h)
        b_pad = f"'Individual Refineries'!$C$3:$C${self.box_last}"
        b_uid = f"'Individual Refineries'!$D$3:$D${self.box_last}"
        b_typ = f"'Individual Refineries'!$A$3:$A${self.box_last}"
        b_cap = f"'Individual Refineries'!$F$3:$F${self.box_last}"
        b_utl = f"'Individual Refineries'!${get_column_letter(self.c_util)}$3:" \
                f"${get_column_letter(self.c_util)}${self.box_last}"
        b_net = f"'Individual Refineries'!${get_column_letter(self.c_base)}$3:" \
                f"${get_column_letter(self.c_base)}${self.box_last}"
        uplift = B["light_uplift"]
        cut = B["heavy_cut"]
        spread = "$D$4"   # gasoline - HVN spread cell above
        scn_first = r + 1
        for i, p in enumerate(self.padds):
            r += 1
            _style(ws.cell(row=r, column=1, value=p), bold=True)
            _style(ws.cell(
                row=r, column=2,
                value=(f'=SUMPRODUCT((({b_uid}="CDU")+({b_uid}="CRUDE-EST"))'
                       f"*({b_pad}={p})*{b_cap}*{b_utl})"),
            ), fill=FILL_CALC, fmt="#,##0")
            _style(ws.cell(
                row=r, column=3,
                value=f'=SUMPRODUCT(({b_typ}="TOTAL")*({b_pad}={p})*{b_net})',
            ), fill=FILL_CALC, fmt="0.0")
            _style(ws.cell(row=r, column=4, value=f"=$B{r}*{uplift}/100"),
                   fill=FILL_CALC, fmt="+0.0;-0.0")
            _style(ws.cell(row=r, column=5, value=f"=-$B{r}*{cut}/100"),
                   fill=FILL_CALC, fmt="+0.0;-0.0")
            _style(ws.cell(row=r, column=6, value=f"=$D{r}*1000*{spread}/1000"),
                   fill=FILL_CALC, fmt="0.0")
        scn_last = r
        r += 2
        ws.cell(row=r, column=1, value=(
            "Where blends become economical: switch to max-light when the naphtha "
            "value chain (spreads + open arbs) beats the light-crude premium plus "
            "logistics; the desk sets those inputs on Assumptions."
        )).font = FONT_SMALL
        ws.column_dimensions["A"].width = 52
        for c in "BCDEF":
            ws.column_dimensions[c].width = 16

        # charts: spreads, arbs, scenario deltas — live off the price inputs
        from openpyxl.chart import BarChart, Reference, Series

        def bar(title, series_specs, cats_ref, anchor, height=7, width=11):
            ch = BarChart()
            ch.type = "col"
            ch.title = title
            ch.height, ch.width = height, width
            if len(series_specs) == 1:
                ch.legend = None
            for ref, label, color in series_specs:
                ser = Series(ref, title=label)
                ser.graphicalProperties.solidFill = color
                ch.series.append(ser)
            ch.set_categories(cats_ref)
            _axes(ch, "", "$/bbl" if "bbl" in title else "kbd")
            ws.add_chart(ch, anchor)

        bar("Blend value spreads ($/bbl)",
            [(Reference(ws, min_col=4, min_row=4, max_row=7), None, NAVY)],
            Reference(ws, min_col=1, min_row=4, max_row=7), "H3")
        bar("Arb netbacks ($/bbl, + = open)",
            [(Reference(ws, min_col=4, min_row=10, max_row=12), None, GOLD)],
            Reference(ws, min_col=1, min_row=10, max_row=12), "H17")
        bar("Slate scenarios: net naphtha delta by PADD (kbd)",
            [(Reference(ws, min_col=4, min_row=scn_first, max_row=scn_last),
              "max-light", "2E8677"),
             (Reference(ws, min_col=5, min_row=scn_first, max_row=scn_last),
              "max-heavy", "B03A2E")],
            Reference(ws, min_col=1, min_row=scn_first, max_row=scn_last),
            "N3", height=7, width=13)

    # ---------------------------------------------------------- KitWalk tab

    def _sheet_kitwalk(self) -> None:
        """The naphtha path per refinery: CDU -> SR naphtha -> NHT ->
        reformer -> net naphtha, against the 2024 actual."""
        ws = self.wb.create_sheet("KitWalk")
        _banner(ws, 1, "Kit walk — CDU -> overheads (SR naphtha) -> naphtha "
                       "hydrofiner -> reformer -> net naphtha vs 2024 actual. The "
                       "yield-tuning workbench: adjust Assumptions/box yields until "
                       "delta ~ 0 per refinery.", 13)
        headers = ["refinery_id", "padd", "CDU cap", "CDU util", "CDU runs",
                   "SR naphtha %", "SR naphtha kbd", "NHT cap", "reformer feed",
                   "isom pull", "net naphtha", "2024 actual %", "implied %",
                   "delta (pts)", "flag"]
        for c, h in enumerate(headers, start=1):
            _hdr(ws, 2, c, h)
        b_rid = f"'Individual Refineries'!$B$3:$B${self.box_last}"
        b_uid = f"'Individual Refineries'!$D$3:$D${self.box_last}"
        b_typ = f"'Individual Refineries'!$A$3:$A${self.box_last}"
        b_cap = f"'Individual Refineries'!$F$3:$F${self.box_last}"
        utl_l = get_column_letter(self.c_util)
        ysm_l = get_column_letter(self.c_ysum)
        b_utl = f"'Individual Refineries'!${utl_l}$3:${utl_l}${self.box_last}"
        b_ysm = f"'Individual Refineries'!${ysm_l}$3:${ysm_l}${self.box_last}"
        b_net = f"'Individual Refineries'!${get_column_letter(self.c_base)}$3:" \
                f"${get_column_letter(self.c_base)}${self.box_last}"
        d_id = f"Data!$A$3:$A${self.data_last_row}"
        d_np = f"Data!$N$3:$N${self.data_last_row}"

        def by_unit(rng, rid_cell, uid):
            return f'SUMPRODUCT(({b_rid}={rid_cell})*({b_uid}="{uid}")*{rng})'

        r = 3
        for ref in self._grid_refineries():
            if not ref.units or ref.units[0].unit_id == "CRUDE-EST":
                continue
            rid = f"$A{r}"
            c = _style(ws.cell(row=r, column=1, value=ref.refinery_id))
            self._link_to_box(c, ref.refinery_id)
            _style(ws.cell(row=r, column=2, value=ref.padd))
            _style(ws.cell(row=r, column=3, value=f'={by_unit(b_cap, rid, "CDU")}'),
                   fill=FILL_CALC, fmt="#,##0")
            _style(ws.cell(row=r, column=4, value=f'={by_unit(b_utl, rid, "CDU")}'),
                   fill=FILL_CALC, fmt="0%")
            _style(ws.cell(row=r, column=5, value=f"=$C{r}*$D{r}"),
                   fill=FILL_CALC, fmt="#,##0")
            _style(ws.cell(row=r, column=6, value=f'={by_unit(b_ysm, rid, "CDU")}'),
                   fill=FILL_CALC, fmt="0.0%")
            _style(ws.cell(row=r, column=7, value=f"=$E{r}*$F{r}"),
                   fill=FILL_CALC, fmt="0.0")
            _style(ws.cell(row=r, column=8, value=f'={by_unit(b_cap, rid, "NHT")}'),
                   fill=FILL_CALC, fmt="#,##0")
            _style(ws.cell(
                row=r, column=9,
                value=(f'={by_unit(b_cap, rid, "REF")}*{by_unit(b_utl, rid, "REF")}'
                       f'+{by_unit(b_cap, rid, "CCR")}*{by_unit(b_utl, rid, "CCR")}'),
            ), fill=FILL_CALC, fmt="0.0")
            _style(ws.cell(
                row=r, column=10,
                value=f'={by_unit(b_cap, rid, "ISOM")}*{by_unit(b_utl, rid, "ISOM")}',
            ), fill=FILL_CALC, fmt="0.0")
            _style(ws.cell(
                row=r, column=11,
                value=f'=SUMPRODUCT(({b_typ}="TOTAL")*({b_rid}={rid})*{b_net})',
            ), fill=FILL_CALC, fmt="0.0")
            _style(ws.cell(row=r, column=12,
                           value=f"=SUMIFS({d_np},{d_id},{rid})"),
                   fill=FILL_CALC, fmt="0.00%")
            _style(ws.cell(
                row=r, column=13,
                value=f'=IFERROR($K{r}/(SUMIFS(Data!$G$3:$G${self.data_last_row},'
                      f'{d_id},{rid})*$D{r}),"")',
            ), fill=FILL_CALC, fmt="0.00%")
            _style(ws.cell(row=r, column=14,
                           value=f'=IF($M{r}="","",$M{r}-$L{r})'),
                   fill=FILL_CALC, fmt="+0.00%;-0.00%")
            _style(ws.cell(
                row=r, column=15,
                value=f'=IF(AND($H{r}>0,$G{r}>$H{r}*1.05),"SR > NHT cap","")',
            ), fill=FILL_CALC)
            r += 1
        last = r - 1
        ws.conditional_formatting.add(
            f"N3:N{last}",
            CellIsRule(operator="greaterThan", formula=["0.02"], fill=FILL_FAIL))
        ws.conditional_formatting.add(
            f"N3:N{last}",
            CellIsRule(operator="lessThan", formula=["-0.02"], fill=FILL_FAIL))
        widths = [24, 6, 8, 8, 8, 10, 11, 8, 12, 9, 11, 11, 10, 11, 13]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "C3"


def build_desk_workbook(
    data: ModelData, axis: list[date], out_path: Path, scenarios=None
) -> Path:
    return DeskWorkbook(data, axis, scenarios=scenarios).build(out_path)


def build_simple_workbook(data: ModelData, axis: list[date], out_path: Path) -> Path:
    return SimpleWorkbook(data, axis).build(out_path)
