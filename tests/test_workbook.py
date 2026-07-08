"""Desk workbook builder tests.

Structural checks always run. If LibreOffice (soffice) is installed, the
workbook is also recalculated headlessly and the Excel numbers are compared
against the Python engine — a full formula round-trip.
"""

import shutil
import subprocess
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from naphtha_model.balance import padd_balance
from naphtha_model.config import DATA_DIR
from naphtha_model.engine import build_axis
from naphtha_model.loaders import load_all
from naphtha_model.scenario import load_scenario
from naphtha_model.workbook import build_desk_workbook

AXIS = build_axis(date(2026, 7, 6), 8)

EXPECTED_SHEETS = [
    "README", "Dashboard", "Boxes", "Balance", "Calibration", "Assumptions",
    "Outages", "Flows", "Demand", "Refineries", "Yields_2024", "Intel",
    "Checks", "Model",
]


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    data = load_all()
    scenarios = [load_scenario(p) for p in sorted((DATA_DIR / "scenarios").glob("*.yaml"))]
    out = tmp_path_factory.mktemp("wb") / "desk.xlsx"
    build_desk_workbook(data, AXIS, out, scenarios=scenarios)
    return data, out


def test_sheets_and_order(built):
    _, out = built
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == EXPECTED_SHEETS


def test_boxes_are_formula_driven(built):
    data, out = built
    wb = openpyxl.load_workbook(out)
    ws = wb["Boxes"]
    unit_rows = [r for r in range(4, 1300) if ws.cell(row=r, column=1).value == "UNIT"]
    assert len(unit_rows) == sum(len(r.units) for r in data.refineries)
    r = unit_rows[0]
    assert str(ws.cell(row=r, column=10).value).startswith("=IF(")     # util
    assert "SUMPRODUCT" in str(ws.cell(row=r, column=25).value)       # offline wk1
    total_rows = [r for r in range(4, 1300) if ws.cell(row=r, column=1).value == "TOTAL"]
    assert len(total_rows) == len(data.refineries)


def test_scenario_rows_and_toggle(built):
    _, out = built
    wb = openpyxl.load_workbook(out)
    ws = wb["Outages"]
    cats = [ws.cell(row=r, column=1).value for r in range(2, 50)]
    assert "scenario" in cats
    scn_row = 2 + cats.index("scenario")
    assert "Model!$B$5" in str(ws.cell(row=scn_row, column=8).value)
    assert wb["Model"]["B5"].value == "NO"
    # scenario flows obey the toggle too
    flows = wb["Flows"]
    assert "Model!$B$5" in str(flows.cell(row=2, column=11).value)


def test_dashboard_charts(built):
    _, out = built
    wb = openpyxl.load_workbook(out)
    assert len(wb["Dashboard"]._charts) >= 5


def test_checks_master_cell(built):
    _, out = built
    wb = openpyxl.load_workbook(out)
    ws = wb["Checks"]
    formulas = [str(ws.cell(row=r, column=3).value) for r in range(4, 20)]
    assert any("ALL CHECKS PASS" in f for f in formulas)


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice not installed")
def test_excel_formulas_match_engine(built, tmp_path):
    """Recalculate with LibreOffice and compare PADD 3 balance to the engine."""
    data, out = built
    profile = tmp_path / "loprofile"
    subprocess.run(
        ["soffice", f"-env:UserInstallation=file://{profile}", "--headless",
         "--convert-to", "xlsx", "--outdir", str(tmp_path), str(out)],
        check=True, capture_output=True, timeout=300,
    )
    recalc = tmp_path / Path(out).name
    wb = openpyxl.load_workbook(recalc, data_only=True)
    bal = wb["Balance"]
    # locate the PADD 3 block: banner, week hdr, supply, base, risk, flows,
    # demand, balance
    banner = next(
        r for r in range(1, 200)
        if str(bal.cell(row=r, column=1).value or "").startswith("PADD 3 ")
    )
    supply_row, risk_row, balance_row = banner + 2, banner + 4, banner + 7
    weeks = padd_balance(
        3, data.refineries, AXIS, data.book, data.outages, data.flows, data.demand
    )
    for k, w in enumerate(weeks):
        col = 2 + k
        assert bal.cell(row=supply_row, column=col).value == pytest.approx(w.supply_kbd, abs=0.05)
        assert bal.cell(row=risk_row, column=col).value == pytest.approx(w.at_risk_kbd, abs=0.05)
        assert bal.cell(row=balance_row, column=col).value == pytest.approx(w.balance_kbd, abs=0.05)
    # every data check must pass
    checks = wb["Checks"]
    statuses = [checks.cell(row=r, column=3).value for r in range(4, 20)]
    assert "FAIL" not in statuses
    assert "ALL CHECKS PASS" in statuses
