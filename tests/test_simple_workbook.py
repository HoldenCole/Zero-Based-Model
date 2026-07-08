"""The desk workbook (default export): one tab per methodology step."""

import shutil
import subprocess
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from naphtha_model.engine import build_axis, refinery_day
from naphtha_model.loaders import load_all
from naphtha_model.workbook import build_simple_workbook

AXIS = build_axis(date(2026, 7, 6), 8)


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    data = load_all()
    out = tmp_path_factory.mktemp("wb") / "simple.xlsx"
    build_simple_workbook(data, AXIS, out)
    return data, out


def _totals(path) -> dict[str, float]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Boxes"]
    return {
        ws.cell(row=r, column=2).value: ws.cell(row=r, column=15).value
        for r in range(3, 3000)
        if ws.cell(row=r, column=1).value == "TOTAL"
    }


def test_tab_layout(built):
    _, out = built
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Data", "Boxes", "Assumptions", "Nameplate",
                             "Effective", "CrudeSlate", "BlendEcon", "KitWalk"]


def test_new_tabs_are_wired(built):
    data, out = built
    wb = openpyxl.load_workbook(out)
    # Nameplate pivots live off Boxes
    assert str(wb["Nameplate"].cell(row=3, column=5).value).startswith("=SUMIFS(Boxes!")
    # KitWalk has one row per unit-detail refinery, formulas over Boxes/Data
    kw = wb["KitWalk"]
    rids = [kw.cell(row=r, column=1).value for r in range(3, 300)
            if kw.cell(row=r, column=1).value]
    unit_detail = [r for r in data.refineries
                   if r.units and r.units[0].unit_id != "CRUDE-EST"]
    assert len(rids) == len(unit_detail)
    assert "SUMPRODUCT" in str(kw.cell(row=3, column=3).value)
    # BlendEcon reads the Assumptions price/freight inputs
    be = wb["BlendEcon"]
    formulas = " ".join(str(be.cell(row=r, column=4).value) for r in range(4, 14))
    assert "Assumptions!" in formulas
    # CrudeSlate carries slate rows and flags naphtha buyers
    cs = wb["CrudeSlate"]
    col8 = [cs.cell(row=r, column=8).value for r in range(3, 2500)]
    assert "BUYER" in col8


def test_every_refinery_has_a_box(built):
    data, out = built
    wb = openpyxl.load_workbook(out)
    ws = wb["Boxes"]
    total_rows = [r for r in range(3, 3000) if ws.cell(row=r, column=1).value == "TOTAL"]
    assert len(total_rows) == len(data.refineries)
    # yield-mode capacity and yields read the Data sheet live
    unit_row = next(
        r for r in range(3, 3000)
        if ws.cell(row=r, column=1).value == "UNIT"
        and ws.cell(row=r, column=4).value == "CRUDE-EST"
    )
    assert "Data!" in str(ws.cell(row=unit_row, column=6).value)
    assert "Data!" in str(ws.cell(row=unit_row, column=13).value)


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice not installed")
def test_boxes_match_engine_and_capacity_lights_up(built, tmp_path):
    data, out = built
    # simulate the capacity sheet: type 500 kbd into a yield-mode refinery
    wb = openpyxl.load_workbook(out)
    ws = wb["Data"]
    row = next(r for r in range(3, 300)
               if ws.cell(row=r, column=1).value == "PBF_DELAWARE_CITY")
    ws.cell(row=row, column=7, value=500)
    capped = tmp_path / "capped.xlsx"
    wb.save(capped)

    profile = tmp_path / "loprofile"
    for f in (out, capped):
        subprocess.run(
            ["soffice", f"-env:UserInstallation=file://{profile}", "--headless",
             "--convert-to", "xlsx", "--outdir", str(tmp_path / "recalc"), str(f)],
            check=True, capture_output=True, timeout=300,
        )
    base = _totals(tmp_path / "recalc" / Path(out).name)
    cap = _totals(tmp_path / "recalc" / "capped.xlsx")

    day = AXIS[0]
    for rid in ("MOTIVA_PAR", "XOM_BAYTOWN", "MPC_GALV_BAY"):
        py = refinery_day(data.refinery(rid), day, data.book, [], include_outages=False).net_kbd
        assert base[rid] == pytest.approx(py, abs=0.05)

    r = data.refinery("PBF_DELAWARE_CITY")
    util = data.book.utilization(r, r.units[0], day).value
    assert base["PBF_DELAWARE_CITY"] == pytest.approx(
        r.crude_capacity_kbd * util * r.naphtha_yield_pct / 100, abs=0.01
    )
    assert cap["PBF_DELAWARE_CITY"] == pytest.approx(
        500 * util * r.naphtha_yield_pct / 100, abs=0.01
    )


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice not installed")
def test_everything_propagates(built, tmp_path):
    """The dynamism guarantee: master dials rescale every box, and an edit on
    the Effective tab flows into the Boxes eff-cap column."""
    data, out = built
    wb = openpyxl.load_workbook(out)
    a = wb["Assumptions"]
    dial = {}
    for row in a.iter_rows(min_row=1, max_row=100, min_col=10, max_col=30):
        for c in row:
            if c.value == "Utilization scaler (100% = as-is)":
                dial["util"] = (c.row, c.column + 1)
            if c.value == "Naphtha yield scaler (100% = as-is)":
                dial["yield"] = (c.row, c.column + 1)
    a.cell(row=dial["util"][0], column=dial["util"][1], value=0.9)
    a.cell(row=dial["yield"][0], column=dial["yield"][1], value=1.2)
    eff = wb["Effective"]
    hdr = {eff.cell(row=2, column=c).value: c for c in range(1, 25)}
    for r in range(3, 200):
        if eff.cell(row=r, column=1).value == "MOTIVA_PAR":
            eff.cell(row=r, column=hdr["CDU"], value=999)
            break
    modded = tmp_path / "modded.xlsx"
    wb.save(modded)

    profile = tmp_path / "loprofile"
    for f in (out, modded):
        subprocess.run(
            ["soffice", f"-env:UserInstallation=file://{profile}", "--headless",
             "--convert-to", "xlsx", "--outdir", str(tmp_path / "recalc"), str(f)],
            check=True, capture_output=True, timeout=300,
        )
    base = _totals(tmp_path / "recalc" / Path(out).name)
    mod = _totals(tmp_path / "recalc" / "modded.xlsx")
    for rid in ("XOM_BAYTOWN", "CHEVRON_PASCAGOULA", "BASF_PORT_ARTHUR"):
        assert mod[rid] == pytest.approx(base[rid] * 0.9 * 1.2, abs=0.05)

    boxes = openpyxl.load_workbook(
        tmp_path / "recalc" / "modded.xlsx", data_only=True)["Boxes"]
    eff_cap = next(
        boxes.cell(row=r, column=7).value for r in range(3, 3000)
        if boxes.cell(row=r, column=1).value == "UNIT"
        and boxes.cell(row=r, column=2).value == "MOTIVA_PAR"
        and boxes.cell(row=r, column=4).value == "CDU"
    )
    assert eff_cap == 999
